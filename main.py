import base64
import json
import os
import re
import uuid
import time
from datetime import datetime
from typing import Optional, Any, List

import anthropic
import pandas as pd
from fastapi import FastAPI, File, UploadFile, HTTPException, Depends
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from config import Settings, get_settings
import asyncio
import shutil
from pathlib import Path

# =========================================================
# APP
# =========================================================

settings = get_settings()

app = FastAPI(
    title=settings.app_title,
    version=settings.app_version,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # testing only; restrict this in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# =========================================================
# CONSTANTS
# =========================================================

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

INPUT_DIR = "input"
os.makedirs(INPUT_DIR, exist_ok=True)

UPLOADED_REQUIREMENT_PATH = os.path.join(INPUT_DIR, "uploaded_requirement.xlsx")
BULK_JOB_DIR = "bulk_jobs"
os.makedirs(BULK_JOB_DIR, exist_ok=True)

ACTIVE_STATUS_PATH = os.path.join(BULK_JOB_DIR, "status.json")
ACTIVE_UPLOAD_DIR = os.path.join(BULK_JOB_DIR, "resumes")

POLLING_OUTPUT_PREFIX = "resume_requirement_output"
OUTPUT_COLUMNS = [
    "Request-ID",
    "MSP Owner",
    "Job Title",
    "Skills - Name",
    "Skills - Experience",
    "Additional Skills",
    "Job Description",
    "Work Location CDF",
    "Rate Card",
    "Annually",
    "Candidate Name",
    "Candidate Phone",
    "Candidate Email",
    "Candidate Location",
    "Candidate Total Experience",
    "Candidate Skills",
    "Experience Mismatch",
    "Skill Mismatch",
    "ATS",
    "Remark",
]

REQUIRED_REQUIREMENT_COLUMNS = [
    "Request-ID",
    "MSP Owner",
    "Job Title",
    "Skills - Name",
    "Skills - Experience",
    "Additional Skills",
    "Job Description",
    "Status",
    "Work Location CDF",
    "Rate Card",
    "Yearly Rate",
]

TRACKER_COLUMNS = [
    "Date",
    "Request ID",
    "Status",
    "Skills",
    "Candidate",
    "Verdict",
    "Call Status",
    "Remarks",
]

PIVOT_REMARK_ORDER = [
    "Exp mismatch",
    "High CTC",
    "High Notice Period",
    "Location Mismatch",
    "Match",
    "Not Suitable",
    "NPU",
    "Skill mismatch",
]

VALID_REMARKS = set(PIVOT_REMARK_ORDER)

VERDICT_COLORS = {
    "Strong Fit": "00B050",
    "Good Fit": "92D050",
    "Possible Fit": "FFC000",
    "Not Suitable": "FF6666",
}


# =========================================================
# RESPONSE MODELS
# =========================================================

class JobSummary(BaseModel):
    request_id: str
    msp_owner: Optional[str] = None
    job_title: Optional[str] = None
    skills_name: Optional[str] = None
    skills_experience: Optional[str] = None
    status: Optional[str] = None
    work_location_cdf: Optional[str] = None
    rate_card: Optional[Any] = None
    annually: Optional[Any] = None


# class BulkAnalyzeResponse(BaseModel):
#     message: str
#     total_files_received: int
#     total_pdf_processed: int
#     total_output_rows: int
#     processing_time_seconds: float
#     output_filename: str
#     download_url: str
#     skipped_files: List[str]
    
    
class BulkAnalyzeResponse(BaseModel):
    message: str
    total_files_received: int
    total_pdf_processed: int
    total_output_rows: int
    processing_time_seconds: float
    output_filename: str
    download_url: str
    skipped_files: List[str]
    token_usage: dict


# =========================================================
# DEBUG
# =========================================================

def debug(msg: str):
    print(f"[DEBUG] {msg}", flush=True)


# =========================================================
# BASIC CLEANING HELPERS
# =========================================================

def clean_cell(value) -> str:
    if value is None:
        return ""

    if pd.isna(value):
        return ""

    value = str(value)
    value = value.replace("_x000D_", "\n")
    value = value.replace("\r", "\n")
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def normalize_text(value) -> str:
    value = clean_cell(value).lower()
    value = value.replace("reactjs", "react js")
    value = value.replace("nodejs", "node js")
    value = value.replace("postgresql", "postgre sql")
    value = value.replace("snowflakes", "snowflake")
    value = value.replace("powerplatform", "power platform")
    value = re.sub(r"[^a-z0-9+#.\s]", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def split_skills(value) -> List[str]:
    value = clean_cell(value)
    parts = re.split(r"[,~|;/\n]+", value)
    return [p.strip() for p in parts if p.strip()]


def safe_number(value, default=0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def is_pdf_upload(file: UploadFile) -> bool:
    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()

    return filename.endswith(".pdf") or content_type == "application/pdf"


def clean_json_text(raw: str) -> dict:
    clean = raw.strip()

    if clean.startswith("```json"):
        clean = clean.removeprefix("```json").strip()

    if clean.startswith("```"):
        clean = clean.removeprefix("```").strip()

    if clean.endswith("```"):
        clean = clean.removesuffix("```").strip()

    start = clean.find("{")
    end = clean.rfind("}")

    if start != -1 and end != -1:
        clean = clean[start:end + 1]

    return json.loads(clean)


# =========================================================
# EXCEL LOADING
# =========================================================
def load_requirement_df(cfg: Settings) -> pd.DataFrame:
    excel_path = UPLOADED_REQUIREMENT_PATH

    debug(f"Loading uploaded requirement Excel: {excel_path}")

    if not os.path.exists(excel_path):
        raise HTTPException(
            status_code=400,
            detail="Requirement Excel not uploaded. Please upload Requirement Excel first.",
        )

    xls = pd.ExcelFile(excel_path)
    debug(f"Available sheets: {xls.sheet_names}")

    matched_sheets = [
        s for s in xls.sheet_names
        if "requirement" in str(s).strip().lower()
    ]

    if matched_sheets:
        sheet_name = matched_sheets[0]
    else:
        sheet_name = xls.sheet_names[0]

    debug(f"Using requirement sheet: {sheet_name}")

    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    rename_map = {
        "Monthly Company Pay Rate": "Rate Card",
        "Monthly Company Pay Rate ": "Rate Card",
        "Anually Company Pay Rate": "Yearly Rate",
        "Annually Company Pay Rate": "Yearly Rate",
        "Annual Company Pay Rate": "Yearly Rate",
        "Request ID": "Request-ID",
        "Request Id": "Request-ID",
        "Work Location City": "Work Location City",
        "Work Location CDF": "Work Location CDF",
    }

    df = df.rename(columns=rename_map)

    missing_cols = [
        col for col in REQUIRED_REQUIREMENT_COLUMNS
        if col not in df.columns
    ]

    if missing_cols:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Invalid Requirement Excel format.",
                "missing_columns": missing_cols,
                "expected_columns": REQUIRED_REQUIREMENT_COLUMNS,
            },
        )

    df["Request-ID"] = df["Request-ID"].astype(str).str.strip()

    df = df[df["Request-ID"].astype(str).str.strip() != ""]
    df = df.dropna(how="all")

    optional_cols = [
        "Work Location City",
        "System Enhancements Required",
        "Candidate Annual CTC",
    ]

    for col in optional_cols:
        if col not in df.columns:
            df[col] = ""

    if df.empty:
        raise HTTPException(
            status_code=400,
            detail="Requirement Excel has no valid rows with Request-ID.",
        )

    debug(f"Requirement dataframe shape: {df.shape}")
    debug(f"Requirement columns: {list(df.columns)}")

    return df


def get_requirement_by_id(request_id: str, cfg: Settings) -> dict:
    df = load_requirement_df(cfg)

    request_id = str(request_id).strip()
    matched = df[df["Request-ID"] == request_id]

    if matched.empty:
        raise HTTPException(
            status_code=404,
            detail=f"Request-ID not found in requirement Excel: {request_id}",
        )

    return matched.iloc[0].fillna("").to_dict()


# =========================================================
# CLAUDE CLIENT
# =========================================================

def get_claude_client(cfg: Settings) -> anthropic.Anthropic:
    return anthropic.Anthropic(api_key=cfg.anthropic_api_key)


# =========================================================
# CLAUDE - RESUME EXTRACTION
# =========================================================

def call_claude_extract_candidate(
    pdf_bytes: bytes,
    filename: str,
    cfg: Settings,
) -> dict:
    debug(f"Claude resume extraction started: {filename}")

    client = get_claude_client(cfg)

    pdf_base64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

    prompt = """
You are an expert resume parser for recruitment screening.

Read the uploaded resume PDF and extract candidate details.

Return only valid JSON. No markdown. No explanation.

Required JSON:
{
  "candidate_name": null,
  "candidate_phone": null,
  "candidate_email": null,
  "candidate_location": null,
  "candidate_total_experience_years": null,
  "candidate_current_ctc": null,
  "candidate_expected_ctc": null,
  "candidate_notice_period": null,
  "candidate_skills": [],
  "recent_job_title": null,
  "profile_summary": null
}

Rules:
- Do not hallucinate missing values.
- If value is missing, return null.
- candidate_total_experience_years should be a number if possible.
- candidate_skills should include programming languages, tools, frameworks, databases, cloud, platforms, domain skills and business tools.
- Extract from resume only.
"""

    message = client.messages.create(
        model=cfg.claude_model,
        max_tokens=cfg.max_tokens,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_base64,
                        },
                    },
                    {
                        "type": "text",
                        "text": prompt,
                    },
                ],
            }
        ],
    )

    raw = message.content[0].text.strip()
    usage = getattr(message, "usage", None)
    input_tokens = getattr(usage, "input_tokens", 0) if usage else 0
    output_tokens = getattr(usage, "output_tokens", 0) if usage else 0

    try:
        result = clean_json_text(raw)
    except Exception as e:
        debug(f"Candidate JSON parse failed for {filename}: {repr(e)}")
        result = {
            "candidate_name": None,
            "candidate_phone": None,
            "candidate_email": None,
            "candidate_location": None,
            "candidate_total_experience_years": None,
            "candidate_current_ctc": None,
            "candidate_expected_ctc": None,
            "candidate_notice_period": None,
            "candidate_skills": [],
            "recent_job_title": None,
            "profile_summary": raw,
        }

    if not isinstance(result.get("candidate_skills"), list):
        result["candidate_skills"] = []

    debug(f"Claude resume extraction completed: {filename}")
    
    result["_token_usage"] = {
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": input_tokens + output_tokens,
    }
    
    return result


# =========================================================
# PYTHON SHORTLISTING OF REQUIREMENTS
# =========================================================

def build_requirement_search_text(row: dict) -> str:
    return " ".join([
        clean_cell(row.get("Job Title")),
        clean_cell(row.get("Skills - Name")),
        clean_cell(row.get("Skills - Experience")),
        clean_cell(row.get("Additional Skills")),
        clean_cell(row.get("Job Description")),
        clean_cell(row.get("Work Location CDF")),
    ])


def build_candidate_search_text(candidate_info: dict) -> str:
    return " ".join([
        " ".join(candidate_info.get("candidate_skills") or []),
        clean_cell(candidate_info.get("recent_job_title")),
        clean_cell(candidate_info.get("profile_summary")),
        clean_cell(candidate_info.get("candidate_location")),
    ])


def shortlist_requirements(
    requirements_df: pd.DataFrame,
    candidate_info: dict,
    top_n: int = 25,
) -> list:
    candidate_text = build_candidate_search_text(candidate_info)
    candidate_norm = normalize_text(candidate_text)
    candidate_tokens = set(candidate_norm.split())

    scored_rows = []

    for _, row in requirements_df.iterrows():
        row_dict = row.fillna("").to_dict()

        req_text = build_requirement_search_text(row_dict)
        req_norm = normalize_text(req_text)
        req_tokens = set(req_norm.split())

        overlap = candidate_tokens.intersection(req_tokens)
        score = len(overlap)

        # Strong boost on exact primary skill tokens
        primary_skills = split_skills(row_dict.get("Skills - Name"))
        for skill in primary_skills:
            skill_norm = normalize_text(skill)
            if skill_norm and skill_norm in candidate_norm:
                score += 25

        # Boost additional skills
        additional_skills = split_skills(row_dict.get("Additional Skills"))
        for skill in additional_skills[:80]:
            skill_norm = normalize_text(skill)
            if skill_norm and skill_norm in candidate_norm:
                score += 5

        # Prefer open requirements slightly
        status = normalize_text(row_dict.get("Status"))
        if status == "open":
            score += 5

        scored_rows.append((score, row_dict))

    scored_rows = sorted(scored_rows, key=lambda x: x[0], reverse=True)

    top_rows = [row for score, row in scored_rows[:top_n]]

    debug(f"Shortlisted {len(top_rows)} requirements")
    return top_rows


# =========================================================
# CLAUDE - BEST REQUIREMENT MATCHING
# =========================================================

def requirement_for_prompt(row: dict) -> dict:
    return {
        "Request-ID": clean_cell(row.get("Request-ID")),
        "MSP Owner": clean_cell(row.get("MSP Owner")),
        "Job Title": clean_cell(row.get("Job Title")),
        "Skills - Name": clean_cell(row.get("Skills - Name")),
        "Skills - Experience": clean_cell(row.get("Skills - Experience")),
        "Additional Skills": clean_cell(row.get("Additional Skills")),
        "Job Description": clean_cell(row.get("Job Description")),
        "Status": clean_cell(row.get("Status")),
        "Work Location City": clean_cell(row.get("Work Location City")),
        "Work Location CDF": clean_cell(row.get("Work Location CDF")),
        "Rate Card": clean_cell(row.get("Rate Card")),
        "Yearly Rate": clean_cell(row.get("Yearly Rate")),
        "System Enhancements Required": clean_cell(row.get("System Enhancements Required")),
        "Candidate Annual CTC": clean_cell(row.get("Candidate Annual CTC")),
    }


def call_claude_best_requirement_match(
    candidate_info: dict,
    shortlisted_requirements: list,
    cfg: Settings,
) -> dict:
    debug("Claude best requirement matching started")

    client = get_claude_client(cfg)

    requirements_for_prompt = [
        requirement_for_prompt(row)
        for row in shortlisted_requirements
    ]

    prompt = f"""
You are an expert technical recruiter.

You are given:
1. One candidate profile extracted from resume.
2. A shortlisted list of job requirements from Requirement Sheet.

Your task:
- Select the best matching Request-ID/job requirement for the candidate.
- If no requirement is a good fit, select the closest Request-ID but mark final_remark as "Not Suitable".
- Generate ATS score from 0 to 100.
- Generate recruiter-style verdict, call_status and remark.

Important:
- Use ONLY the provided requirements.
- Do not invent Request-ID.
- NPU means Not Picked Up. It cannot be detected from resume. Do not set NPU unless input clearly says call status is NPU.
- Use High CTC only if candidate CTC is available and exceeds requirement budget.
- Use High Notice Period only if notice period is available and clearly high.
- Use Location Mismatch if location is main issue.
- Use Skill mismatch if skill gap is main issue.
- Use Exp mismatch if experience is main issue.
- Use Match only when candidate is a strong/clear fit.
- Use Not Suitable when profile/domain is not aligned.
- call_status must always be empty string "" because call status is recruiter-entered manually after calling the candidate.

Allowed verdict values:
- Strong Fit
- Good Fit
- Possible Fit
- Not Suitable

Allowed final_remark values:
- Exp mismatch
- High CTC
- High Notice Period
- Location Mismatch
- Match
- Not Suitable
- Skill mismatch

Set experience_mismatch as "Yes" if candidate total experience is less than required experience, otherwise "No".
Set skill_mismatch as "Yes" if candidate is missing important required skills, otherwise "No".

Candidate Profile:
{json.dumps(candidate_info, ensure_ascii=False, indent=2)}

Shortlisted Requirements:
{json.dumps(requirements_for_prompt, ensure_ascii=False, indent=2)}

Return only valid JSON:
{{
  "best_request_id": null,
  "ats_score": 0,
  "verdict": "Not Suitable",
  "call_status": "",
  "final_remark": "Not Suitable",
  "experience_mismatch": "No",
  "skill_mismatch": "No",
  "matching_skills": [],
  "missing_skills": [],
  "reason": ""
}}
"""

    message = client.messages.create(
        model=cfg.claude_model,
        max_tokens=cfg.max_tokens,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt,
                    }
                ],
            }
        ],
    )

    raw = message.content[0].text.strip()
    
    usage = getattr(message, "usage", None)
    input_tokens = getattr(usage, "input_tokens", 0) if usage else 0
    output_tokens = getattr(usage, "output_tokens", 0) if usage else 0

    try:
        result = clean_json_text(raw)
    except Exception as e:
        debug(f"Match JSON parse failed: {repr(e)}")
        # result = {
        #     "best_request_id": None,
        #     "ats_score": 0,
        #     "verdict": "Not Suitable",
        #     "call_status": "",
        #     "final_remark": "Not Suitable",
        #     "matching_skills": [],
        #     "missing_skills": [],
        #     "reason": raw,
        # }
        result = {
            "best_request_id": None,
            "ats_score": 0,
            "verdict": "Not Suitable",
            "call_status": "",
            "final_remark": "Not Suitable",
            "experience_mismatch": "No",
            "skill_mismatch": "No",
            "matching_skills": [],
            "missing_skills": [],
            "reason": raw,
        }

    final_remark = clean_cell(result.get("final_remark"))

    if final_remark not in VALID_REMARKS:
        final_remark = "Not Suitable"

    result["final_remark"] = final_remark

    verdict = clean_cell(result.get("verdict"))
    if verdict not in ["Strong Fit", "Good Fit", "Possible Fit", "Not Suitable"]:
        ats = safe_number(result.get("ats_score"))
        if ats >= 80:
            verdict = "Strong Fit"
        elif ats >= 60:
            verdict = "Good Fit"
        elif ats >= 40:
            verdict = "Possible Fit"
        else:
            verdict = "Not Suitable"

    result["verdict"] = verdict

    debug("Claude best requirement matching completed")
    
    result["_token_usage"] = {
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": input_tokens + output_tokens,
    }
    
    return result


# =========================================================
# OUTPUT ROW BUILDERS
# =========================================================

def find_requirement_by_request_id(requirements_df: pd.DataFrame, request_id: str) -> dict:
    request_id = clean_cell(request_id)

    if not request_id:
        return {}

    matched = requirements_df[
        requirements_df["Request-ID"].astype(str).str.strip() == request_id
    ]

    if matched.empty:
        return {}

    return matched.iloc[0].fillna("").to_dict()


def build_output_row(req: dict, candidate_info: dict, match_result: dict) -> dict:
    candidate_skills = candidate_info.get("candidate_skills") or []

    return {
        "Request-ID": clean_cell(req.get("Request-ID")),
        "MSP Owner": clean_cell(req.get("MSP Owner")),
        "Job Title": clean_cell(req.get("Job Title")),
        "Skills - Name": clean_cell(req.get("Skills - Name")),
        "Skills - Experience": clean_cell(req.get("Skills - Experience")),
        "Additional Skills": clean_cell(req.get("Additional Skills")),
        "Job Description": clean_cell(req.get("Job Description")),
        "Work Location CDF": clean_cell(req.get("Work Location CDF")),
        "Rate Card": clean_cell(req.get("Rate Card")),
        "Annually": clean_cell(req.get("Yearly Rate")),
        "Candidate Name": clean_cell(candidate_info.get("candidate_name")),
        "Candidate Phone": clean_cell(candidate_info.get("candidate_phone")),
        "Candidate Email": clean_cell(candidate_info.get("candidate_email")),
        "Candidate Location": clean_cell(candidate_info.get("candidate_location")),
        "Candidate Total Experience": clean_cell(candidate_info.get("candidate_total_experience_years")),
        "Candidate Skills": ", ".join(candidate_skills),
        "Experience Mismatch": clean_cell(match_result.get("experience_mismatch")) or "No",
        "Skill Mismatch": clean_cell(match_result.get("skill_mismatch")) or "No",
        "ATS": clean_cell(match_result.get("ats_score")),
        "Remark": clean_cell(match_result.get("reason")),
    }


def build_tracker_row(req: dict, candidate_info: dict, match_result: dict, filename: str) -> dict:
    candidate_name = clean_cell(candidate_info.get("candidate_name"))

    if not candidate_name:
        candidate_name = filename

    return {
        "Date": datetime.now().strftime("%d-%m-%Y"),
        "Request ID": clean_cell(req.get("Request-ID")) or clean_cell(match_result.get("best_request_id")),
        "Status": clean_cell(req.get("Status")),
        "Skills": clean_cell(req.get("Skills - Name")),
        "Candidate": candidate_name,
        "Verdict": clean_cell(match_result.get("verdict")),
        "Call Status": "",
        "Remarks": clean_cell(match_result.get("final_remark")),
    }


# =========================================================
# EXCEL GENERATION
# =========================================================

def style_output_sheet(ws):
    header_blue = PatternFill("solid", fgColor="B7DEE8")
    candidate_green = PatternFill("solid", fgColor="DAF2D0")
    result_pink = PatternFill("solid", fgColor="F2CEEF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        if cell.column <= 10:
            cell.fill = header_blue
        elif 11 <= cell.column <= 16:
            cell.fill = candidate_green
        else:
            cell.fill = result_pink

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_tracker_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    verdict_col = None
    for cell in ws[1]:
        if cell.value == "Verdict":
            verdict_col = cell.column
            break

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

        if verdict_col:
            verdict_cell = row[verdict_col - 1]
            verdict = clean_cell(verdict_cell.value)
            color = VERDICT_COLORS.get(verdict)

            if color:
                verdict_cell.fill = PatternFill("solid", fgColor=color)
                verdict_cell.font = Font(bold=True, color="FFFFFF")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_pivot_sheet(ws):
    header_fill = PatternFill("solid", fgColor="B7DEE8")
    total_fill = PatternFill("solid", fgColor="B7DEE8")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

        first_cell = row[0]
        if clean_cell(first_cell.value) == "Grand Total":
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = total_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_adjust_columns(ws):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0

        for cell in col_cells:
            value = clean_cell(cell.value)
            if len(value) > max_len:
                max_len = len(value[:100])

        width = min(max(max_len + 2, 12), 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_pivot_df(tracker_df: pd.DataFrame) -> pd.DataFrame:
    counts = tracker_df["Remarks"].fillna("Not Suitable").value_counts().to_dict()

    rows = []

    for remark in PIVOT_REMARK_ORDER:
        count = int(counts.get(remark, 0))
        if count > 0:
            rows.append({
                "Row Labels": remark,
                "Count of Remarks": count,
            })

    grand_total = sum(row["Count of Remarks"] for row in rows)

    rows.append({
        "Row Labels": "Grand Total",
        "Count of Remarks": grand_total,
    })

    return pd.DataFrame(rows, columns=["Row Labels", "Count of Remarks"])


def create_final_excel(output_rows: list, tracker_rows: list, output_path: str):
    output_df = pd.DataFrame(output_rows, columns=OUTPUT_COLUMNS)
    tracker_df = pd.DataFrame(tracker_rows, columns=TRACKER_COLUMNS)
    pivot_df = create_pivot_df(tracker_df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        output_df.to_excel(writer, sheet_name="Output Sheet", index=False)
        tracker_df.to_excel(writer, sheet_name="Tracker", index=False)
        pivot_df.to_excel(writer, sheet_name="Pivot", index=False)

        workbook = writer.book

        ws_output = workbook["Output Sheet"]
        ws_tracker = workbook["Tracker"]
        ws_pivot = workbook["Pivot"]

        style_output_sheet(ws_output)
        style_tracker_sheet(ws_tracker)
        style_pivot_sheet(ws_pivot)

        auto_adjust_columns(ws_output)
        auto_adjust_columns(ws_tracker)
        auto_adjust_columns(ws_pivot)

    debug(f"Final Excel created: {output_path}")
    
    
    
    
    
def get_default_status():
    return {
        "job_id": "",
        "status": "idle",  # idle, queued, processing, completed, failed
        "message": "No active job",
        "total": 0,
        "processed": 0,
        "successful": 0,
        "failed": 0,
        "skipped": 0,
        "current_file": "",
        "current_batch": "",
        "output_filename": "",
        "download_url": "",
        "failed_files": [],
        "skipped_files": [],
        "token_usage": {
            "input_tokens": 0,
            "output_tokens": 0,
            "total_tokens": 0,
        },
        "started_at": "",
        "updated_at": "",
        "completed_at": "",
    }


def read_status():
    if not os.path.exists(ACTIVE_STATUS_PATH):
        return get_default_status()

    try:
        with open(ACTIVE_STATUS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return get_default_status()


def write_status(data: dict):
    data["updated_at"] = datetime.now().isoformat()

    with open(ACTIVE_STATUS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def is_active_job_running():
    status = read_status()
    return status.get("status") in ["queued", "processing"]


def reset_bulk_job_storage():
    if os.path.exists(ACTIVE_STATUS_PATH):
        os.remove(ACTIVE_STATUS_PATH)

    if os.path.exists(ACTIVE_UPLOAD_DIR):
        shutil.rmtree(ACTIVE_UPLOAD_DIR)

    os.makedirs(ACTIVE_UPLOAD_DIR, exist_ok=True)


class SavedUploadFile:
    def __init__(self, filename: str, content_type: str, path: str):
        self.filename = filename
        self.content_type = content_type
        self.path = path

    async def read(self):
        with open(self.path, "rb") as f:
            return f.read()


# =========================================================
# ROUTES
# =========================================================

@app.get("/")
def root():
    return {
        "message": "Resume Requirement Bulk Matching API",
        "usage": {
            "list_jobs": "GET /jobs",
            "get_job": "GET /jobs/{request_id}",
            "bulk_analyze": "POST /bulk-analyze with multiple PDF resume files",
            "download": "GET /download/{filename}",
        },
    }

@app.post("/upload-requirement")
async def upload_requirement_excel(
    file: UploadFile = File(...),
    cfg: Settings = Depends(get_settings),
):
    filename = file.filename or ""

    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Please upload a valid Excel file with .xlsx or .xls extension.",
        )

    content = await file.read()

    if not content:
        raise HTTPException(status_code=400, detail="Uploaded Excel file is empty.")

    temp_path = os.path.join(INPUT_DIR, f"temp_{uuid.uuid4().hex}_{filename}")

    try:
        with open(temp_path, "wb") as f:
            f.write(content)

        xls = pd.ExcelFile(temp_path)

        matched_sheets = [
            s for s in xls.sheet_names
            if "requirement" in str(s).strip().lower()
        ]

        if matched_sheets:
            sheet_name = matched_sheets[0]
        else:
            sheet_name = xls.sheet_names[0]

        df = pd.read_excel(temp_path, sheet_name=sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        rename_map = {
            "Monthly Company Pay Rate": "Rate Card",
            "Monthly Company Pay Rate ": "Rate Card",
            "Anually Company Pay Rate": "Yearly Rate",
            "Annually Company Pay Rate": "Yearly Rate",
            "Annual Company Pay Rate": "Yearly Rate",
            "Request ID": "Request-ID",
            "Request Id": "Request-ID",
            "Work Location City": "Work Location City",
            "Work Location CDF": "Work Location CDF",
        }

        df = df.rename(columns=rename_map)

        missing_cols = [
            col for col in REQUIRED_REQUIREMENT_COLUMNS
            if col not in df.columns
        ]

        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "Invalid Requirement Excel format.",
                    "missing_columns": missing_cols,
                    "expected_columns": REQUIRED_REQUIREMENT_COLUMNS,
                },
            )

        df["Request-ID"] = df["Request-ID"].astype(str).str.strip()
        df = df[df["Request-ID"].astype(str).str.strip() != ""]
        df = df.dropna(how="all")

        if df.empty:
            raise HTTPException(
                status_code=400,
                detail="Requirement Excel has no valid rows with Request-ID.",
            )

        with open(UPLOADED_REQUIREMENT_PATH, "wb") as f:
            f.write(content)

        return {
            "success": True,
            "message": "Requirement Excel uploaded and validated successfully.",
            "filename": filename,
            "sheet_used": sheet_name,
            "total_jobs": len(df),
        }

    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)
            
            
            
@app.post("/start-bulk-analyze")
async def start_bulk_analyze_resumes(
    files: List[UploadFile] = File(...),
    cfg: Settings = Depends(get_settings),
):
    debug("=" * 100)
    debug("Start bulk analyze request received")
    debug(f"Total files received: {len(files)}")

    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    if not os.path.exists(UPLOADED_REQUIREMENT_PATH):
        raise HTTPException(
            status_code=400,
            detail="Please upload valid Requirement Excel before analyzing resumes.",
        )

    if is_active_job_running():
        raise HTTPException(
            status_code=409,
            detail="A resume analysis job is already running. Please wait until it completes.",
        )

    # Delete old status and old uploaded resumes before starting new batch
    reset_bulk_job_storage()

    job_id = f"JOB_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{str(uuid.uuid4())[:8]}"

    saved_files = []

    for index, file in enumerate(files, start=1):
        safe_name = os.path.basename(file.filename or f"resume_{index}.pdf")
        file_path = os.path.join(ACTIVE_UPLOAD_DIR, safe_name)

        content = await file.read()

        with open(file_path, "wb") as f:
            f.write(content)

        saved_files.append({
            "filename": safe_name,
            "content_type": file.content_type or "application/pdf",
            "path": file_path,
        })

    status = get_default_status()
    status.update({
        "job_id": job_id,
        "status": "queued",
        "message": "Resume analysis job queued",
        "total": len(saved_files),
        "processed": 0,
        "successful": 0,
        "failed": 0,
        "skipped": 0,
        "current_file": "",
        "current_batch": f"0/{len(saved_files)}",
        "started_at": datetime.now().isoformat(),
    })

    write_status(status)

    asyncio.create_task(run_bulk_analyze_background_job(job_id, saved_files, cfg))

    return {
        "job_id": job_id,
        "status": "queued",
        "total": len(saved_files),
        "message": "Resume analysis started",
    }
            

@app.get("/bulk-status")
def get_bulk_status():
    return read_status()



@app.post("/reset-bulk-status")
def reset_bulk_status():
    if is_active_job_running():
        raise HTTPException(
            status_code=409,
            detail="Cannot reset while resume analysis is running.",
        )

    reset_bulk_job_storage()

    status = get_default_status()
    write_status(status)

    return {
        "success": True,
        "message": "Bulk status reset successfully.",
    }
      
async def run_bulk_analyze_background_job(
    job_id: str,
    saved_files: list,
    cfg: Settings,
):
    start_time = time.time()

    output_rows = []
    tracker_rows = []

    total_input_tokens = 0
    total_output_tokens = 0
    total_tokens = 0

    failed_files = []
    skipped_files = []

    try:
        status = read_status()
        status.update({
            "status": "processing",
            "message": "Processing resumes",
        })
        write_status(status)

        requirements_df = load_requirement_df(cfg)

        total_files = len(saved_files)

        # Keep 1 for now. Later make this 2 or 3 if API rate limit allows.
        for index, item in enumerate(saved_files, start=1):
            filename = item["filename"]

            status = read_status()
            status.update({
                "status": "processing",
                "message": f"Processing resume {index} of {total_files}",
                "current_file": filename,
                "current_batch": f"{index}/{total_files}",
            })
            write_status(status)

            fake_file = SavedUploadFile(
                filename=filename,
                content_type=item.get("content_type") or "application/pdf",
                path=item["path"],
            )

            try:
                result = await process_single_resume_file(
                    file=fake_file,
                    requirements_df=requirements_df,
                    cfg=cfg,
                    index=index,
                    total_files=total_files,
                )

                usage = result.get("token_usage", {})

                total_input_tokens += int(usage.get("input_tokens", 0))
                total_output_tokens += int(usage.get("output_tokens", 0))
                total_tokens += int(usage.get("total_tokens", 0))

                if result["status"] == "skipped":
                    skipped_files.append(result["filename"])
                    status = read_status()
                    status["skipped"] = len(skipped_files)
                    status["processed"] = index
                    status["skipped_files"] = skipped_files
                    status["token_usage"] = {
                        "input_tokens": total_input_tokens,
                        "output_tokens": total_output_tokens,
                        "total_tokens": total_tokens,
                    }
                    write_status(status)
                    continue

                if result.get("output_row"):
                    output_rows.append(result["output_row"])

                if result.get("tracker_row"):
                    tracker_rows.append(result["tracker_row"])

                status = read_status()
                status["processed"] = index
                status["successful"] = len(output_rows)
                status["skipped"] = len(skipped_files)
                status["failed"] = len(failed_files)
                status["token_usage"] = {
                    "input_tokens": total_input_tokens,
                    "output_tokens": total_output_tokens,
                    "total_tokens": total_tokens,
                }
                write_status(status)

            except Exception as e:
                debug(f"Resume failed inside background job: {filename}: {repr(e)}")

                failed_files.append({
                    "filename": filename,
                    "error": str(e),
                })

                status = read_status()
                status["processed"] = index
                status["failed"] = len(failed_files)
                status["failed_files"] = failed_files
                status["message"] = f"Resume failed: {filename}. Continuing next resume."
                write_status(status)

        if not output_rows:
            status = read_status()
            status.update({
                "status": "failed",
                "message": "No valid PDF resumes were processed.",
                "completed_at": datetime.now().isoformat(),
            })
            write_status(status)
            return

        output_filename = f"{POLLING_OUTPUT_PREFIX}_{job_id}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        create_final_excel(
            output_rows=output_rows,
            tracker_rows=tracker_rows,
            output_path=output_path,
        )

        total_time = round(time.time() - start_time, 2)

        status = read_status()
        status.update({
            "status": "completed",
            "message": "Bulk resume analysis completed successfully.",
            "processed": total_files,
            "successful": len(output_rows),
            "failed": len(failed_files),
            "skipped": len(skipped_files),
            "current_file": "",
            "current_batch": f"{total_files}/{total_files}",
            "output_filename": output_filename,
            "download_url": f"/download/{output_filename}",
            "failed_files": failed_files,
            "skipped_files": skipped_files,
            "processing_time_seconds": total_time,
            "token_usage": {
                "input_tokens": total_input_tokens,
                "output_tokens": total_output_tokens,
                "total_tokens": total_tokens,
            },
            "completed_at": datetime.now().isoformat(),
        })
        write_status(status)

    except Exception as e:
        debug(f"Background bulk job failed: {repr(e)}")

        status = read_status()
        status.update({
            "status": "failed",
            "message": str(e),
            "completed_at": datetime.now().isoformat(),
        })
        write_status(status)
            
@app.get("/jobs")
def list_jobs(cfg: Settings = Depends(get_settings)):
    if not os.path.exists(UPLOADED_REQUIREMENT_PATH):
        return {
            "requirement_uploaded": False,
            "total_jobs": 0,
            "jobs": [],
            "message": "Requirement Excel not uploaded.",
        }

    df = load_requirement_df(cfg)

    jobs = []

    for _, row in df.iterrows():
        row = row.fillna("").to_dict()

        jobs.append({
            "request_id": clean_cell(row.get("Request-ID")),
            "msp_owner": clean_cell(row.get("MSP Owner")),
            "job_title": clean_cell(row.get("Job Title")),
            "skills_name": clean_cell(row.get("Skills - Name")),
            "skills_experience": clean_cell(row.get("Skills - Experience")),
            "status": clean_cell(row.get("Status")),
            "work_location_cdf": clean_cell(row.get("Work Location CDF")),
            "rate_card": clean_cell(row.get("Rate Card")),
            "annually": clean_cell(row.get("Yearly Rate")),
        })

    return {
        "requirement_uploaded": True,
        "total_jobs": len(jobs),
        "jobs": jobs,
    }


@app.get("/jobs/{request_id}")
def get_job(request_id: str, cfg: Settings = Depends(get_settings)):
    requirement = get_requirement_by_id(request_id, cfg)

    return {
        "request_id": request_id,
        "requirement": requirement,
    }

async def process_single_resume_file(
    file: UploadFile,
    requirements_df: pd.DataFrame,
    cfg: Settings,
    index: int,
    total_files: int,
):
    debug("-" * 100)
    debug(f"Processing file {index}/{total_files}: {file.filename}")
    debug(f"Content type: {file.content_type}")

    zero_token_usage = {
        "input_tokens": 0,
        "output_tokens": 0,
        "total_tokens": 0,
    }

    if not is_pdf_upload(file):
        return {
            "status": "skipped",
            "filename": file.filename or "unknown_file",
            "output_row": None,
            "tracker_row": None,
            "token_usage": zero_token_usage,
        }

    pdf_bytes = await file.read()

    if not pdf_bytes:
        return {
            "status": "skipped",
            "filename": file.filename or "empty_file",
            "output_row": None,
            "tracker_row": None,
            "token_usage": zero_token_usage,
        }

    if len(pdf_bytes) > cfg.max_pdf_size_mb * 1024 * 1024:
        return {
            "status": "skipped",
            "filename": file.filename or "large_file",
            "output_row": None,
            "tracker_row": None,
            "token_usage": zero_token_usage,
        }

    try:
        candidate_info = await asyncio.to_thread(
            call_claude_extract_candidate,
            pdf_bytes,
            file.filename or "resume.pdf",
            cfg,
        )

        shortlisted_requirements = shortlist_requirements(
            requirements_df=requirements_df,
            candidate_info=candidate_info,
            top_n=10,
        )

        match_result = await asyncio.to_thread(
            call_claude_best_requirement_match,
            candidate_info,
            shortlisted_requirements,
            cfg,
        )

        candidate_usage = candidate_info.get("_token_usage", {})
        match_usage = match_result.get("_token_usage", {})

        token_usage = {
            "input_tokens": int(candidate_usage.get("input_tokens", 0))
            + int(match_usage.get("input_tokens", 0)),

            "output_tokens": int(candidate_usage.get("output_tokens", 0))
            + int(match_usage.get("output_tokens", 0)),

            "total_tokens": int(candidate_usage.get("total_tokens", 0))
            + int(match_usage.get("total_tokens", 0)),
        }

        best_request_id = clean_cell(match_result.get("best_request_id"))

        req = find_requirement_by_request_id(
            requirements_df=requirements_df,
            request_id=best_request_id,
        )

        if not req and shortlisted_requirements:
            req = shortlisted_requirements[0]
            match_result["best_request_id"] = clean_cell(req.get("Request-ID"))
            match_result["final_remark"] = "Not Suitable"
            match_result["verdict"] = "Not Suitable"
            match_result["ats_score"] = 0
            match_result["experience_mismatch"] = "No"
            match_result["skill_mismatch"] = "No"
            match_result["reason"] = (
                "No confident matching Request-ID returned by AI. "
                "Closest requirement selected as fallback, but candidate marked Not Suitable."
            )

        output_row = build_output_row(
            req=req,
            candidate_info=candidate_info,
            match_result=match_result,
        )

        tracker_row = build_tracker_row(
            req=req,
            candidate_info=candidate_info,
            match_result=match_result,
            filename=file.filename or "resume.pdf",
        )

        return {
            "status": "processed",
            "filename": file.filename or "resume.pdf",
            "output_row": output_row,
            "tracker_row": tracker_row,
            "token_usage": token_usage,
        }

    except Exception as e:
        debug(f"Error processing file {file.filename}: {repr(e)}")

        candidate_name = file.filename or "unknown_resume"

        error_output_row = {col: "" for col in OUTPUT_COLUMNS}
        error_output_row["Candidate Name"] = candidate_name
        error_output_row["ATS"] = "0"
        error_output_row["Experience Mismatch"] = "No"
        error_output_row["Skill Mismatch"] = "No"
        error_output_row["Remark"] = f"Error while processing resume: {str(e)}"

        tracker_row = {
            "Date": datetime.now().strftime("%d-%m-%Y"),
            "Request ID": "",
            "Status": "",
            "Skills": "",
            "Candidate": candidate_name,
            "Verdict": "Not Suitable",
            "Call Status": "",
            "Remarks": "Not Suitable",
        }

        return {
            "status": "processed",
            "filename": candidate_name,
            "output_row": error_output_row,
            "tracker_row": tracker_row,
            "token_usage": zero_token_usage,
        }
        
@app.post("/bulk-analyze", response_model=BulkAnalyzeResponse)
async def bulk_analyze_resumes(
    files: List[UploadFile] = File(...),
    cfg: Settings = Depends(get_settings),
):
    start_time = time.time()

    debug("=" * 100)
    debug("Bulk analyze request received")
    debug(f"Total files received: {len(files)}")

    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    
    if not os.path.exists(UPLOADED_REQUIREMENT_PATH):
        raise HTTPException(
            status_code=400,
            detail="Please upload valid Requirement Excel before analyzing resumes.",
        )

    requirements_df = load_requirement_df(cfg)

    if requirements_df.empty:
        raise HTTPException(status_code=400, detail="Requirement sheet has no rows")

    # Limit parallel Claude calls for safety.
    # For demo keep 2 or 3. Do not set very high.
    semaphore = asyncio.Semaphore(1)

    async def limited_process(file, index):
        async with semaphore:
            return await process_single_resume_file(
                file=file,
                requirements_df=requirements_df,
                cfg=cfg,
                index=index,
                total_files=len(files),
            )

    tasks = [
        limited_process(file, index)
        for index, file in enumerate(files, start=1)
    ]

    results = await asyncio.gather(*tasks)

    output_rows = []
    tracker_rows = []
    skipped_files = []
    processed_count = 0

    total_input_tokens = 0
    total_output_tokens = 0
    total_tokens = 0

    for result in results:
        usage = result.get("token_usage", {})

        total_input_tokens += int(usage.get("input_tokens", 0))
        total_output_tokens += int(usage.get("output_tokens", 0))
        total_tokens += int(usage.get("total_tokens", 0))

        if result["status"] == "skipped":
            skipped_files.append(result["filename"])
            continue

        if result["output_row"]:
            output_rows.append(result["output_row"])

        if result["tracker_row"]:
            tracker_rows.append(result["tracker_row"])

        processed_count += 1

    if not output_rows:
        raise HTTPException(
            status_code=400,
            detail="No valid PDF resumes were processed",
        )

    file_id = str(uuid.uuid4())[:8]
    output_filename = f"resume_requirement_output_{file_id}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)

    create_final_excel(
        output_rows=output_rows,
        tracker_rows=tracker_rows,
        output_path=output_path,
    )

    total_time = round(time.time() - start_time, 2)

    return BulkAnalyzeResponse(
        message="Bulk resume analysis completed",
        total_files_received=len(files),
        total_pdf_processed=processed_count,
        total_output_rows=len(output_rows),
        processing_time_seconds=total_time,
        output_filename=output_filename,
        download_url=f"/download/{output_filename}",
        skipped_files=skipped_files,
        token_usage={
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "total_tokens": total_tokens,
        },
    )

@app.get("/download/{filename}")
def download_output_file(filename: str):
    # Prevent path traversal
    filename = os.path.basename(filename)
    file_path = os.path.join(OUTPUT_DIR, filename)

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )