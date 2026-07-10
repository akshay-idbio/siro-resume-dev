import base64
import json
import os
import re
import uuid
import time
from datetime import datetime
from typing import Optional, Any, List

from db import init_db_indexes
from auth import auth_router, admin_router

import anthropic
import pandas as pd

from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, BackgroundTasks
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from config import Settings, get_settings
import asyncio
import glob
from auth import get_current_user
from job_db import (
    get_user_job_by_id,
    mark_job_processing,
    mark_job_completed,
    mark_job_failed,
    mark_resume_processing,
    mark_resume_completed,
    mark_resume_failed,
    update_job_progress,
    save_job_results,
)

from auth import auth_router, admin_router, get_current_user
import shutil
import subprocess
import tempfile
from pathlib import Path
from job_upload_routes import job_upload_router

from job_routes import jobs_router, admin_jobs_router

# =========================================================
# APP
# =========================================================

settings = get_settings()

app = FastAPI(
    title=settings.app_title,
    version=settings.app_version,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # testing only; restrict this in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    await init_db_indexes()


app.include_router(auth_router)
app.include_router(admin_router)

app.include_router(jobs_router)
app.include_router(admin_jobs_router)

app.include_router(job_upload_router)


# =========================================================
# CONSTANTS
# =========================================================

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

INPUT_DIR = "input"
os.makedirs(INPUT_DIR, exist_ok=True)

UPLOADED_REQUIREMENT_PATH = os.path.join(INPUT_DIR, "uploaded_requirement.xlsx")
BULK_JOB_DIR = "bulk_jobs"
os.makedirs(BULK_JOB_DIR, exist_ok=True)

ACTIVE_STATUS_PATH = os.path.join(BULK_JOB_DIR, "status.json")
ACTIVE_UPLOAD_DIR = os.path.join(BULK_JOB_DIR, "resumes")

POLLING_OUTPUT_PREFIX = "resume_requirement_output"
# OUTPUT_COLUMNS = [
#     "Request-ID",
#     "MSP Owner",
#     "Job Title",
#     "Skills - Name",
#     "Skills - Experience",
#     "Additional Skills",
#     "Job Description",
#     "Work Location CDF",
#     "Rate Card",
#     "Annually",
#     "Candidate Name",
#     "Candidate Phone",
#     "Candidate Email",
#     "Candidate Location",
#     "Candidate Total Experience",
#     "Candidate Skills",
#     "Experience Mismatch",
#     "Skill Mismatch",
#     "ATS",
#     "Remark",
# ]

OUTPUT_COLUMNS = [
    "Request-ID",
    "MSP Owner",
    "Job Title",
    "Skills - Name",
    "Skills - Experience",
    "Additional Skills",
    "Job Description",
    "Work Location CDF",
    "Rate Card",
    "Annually",
    "CV File Name",
    "Candidate Name",
    "Candidate Phone",
    "Candidate Email",
    "Candidate Location",
    "Candidate Total Experience",
    "Candidate Skills",
    "Experience Mismatch",
    "Skill Mismatch",
    "ATS",
    "Remark",
]

REQUIRED_REQUIREMENT_COLUMNS = [
    "Request-ID",
    "MSP Owner",
    "Job Title",
    "Skills - Name",
    "Skills - Experience",
    "Additional Skills",
    "Job Description",
    "Status",
    "Work Location CDF",
    "Rate Card",
    "Yearly Rate",
]

TRACKER_COLUMNS = [
    "Date",
    "Request ID",
    "Status",
    "Skills",
    "Candidate",
    "Verdict",
    "Call Status",
    "Remarks",
]

PIVOT_REMARK_ORDER = [
    "Exp mismatch",
    "High CTC",
    "High Notice Period",
    "Location Mismatch",
    "Match",
    "Not Suitable",
    "NPU",
    "Skill mismatch",
]

VALID_REMARKS = set(PIVOT_REMARK_ORDER)

VERDICT_COLORS = {
    "Strong Fit": "00B050",
    "Good Fit": "92D050",
    "Possible Fit": "FFC000",
    "Not Suitable": "FF6666",
}


# =========================================================
# RESPONSE MODELS
# =========================================================


class JobSummary(BaseModel):
    request_id: str
    msp_owner: Optional[str] = None
    job_title: Optional[str] = None
    skills_name: Optional[str] = None
    skills_experience: Optional[str] = None
    status: Optional[str] = None
    work_location_cdf: Optional[str] = None
    rate_card: Optional[Any] = None
    annually: Optional[Any] = None


# class BulkAnalyzeResponse(BaseModel):
#     message: str
#     total_files_received: int
#     total_pdf_processed: int
#     total_output_rows: int
#     processing_time_seconds: float
#     output_filename: str
#     download_url: str
#     skipped_files: List[str]


class BulkAnalyzeResponse(BaseModel):
    message: str
    total_files_received: int
    total_pdf_processed: int
    total_output_rows: int
    processing_time_seconds: float
    output_filename: str
    download_url: str
    skipped_files: List[str]
    token_usage: dict


# =========================================================
# DEBUG
# =========================================================


def debug(msg: str):
    print(f"[DEBUG] {msg}", flush=True)


# =========================================================
# BASIC CLEANING HELPERS
# =========================================================


def clean_cell(value) -> str:
    if value is None:
        return ""

    if pd.isna(value):
        return ""

    value = str(value)
    value = value.replace("_x000D_", "\n")
    value = value.replace("\r", "\n")
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def normalize_text(value) -> str:
    value = clean_cell(value).lower()
    value = value.replace("reactjs", "react js")
    value = value.replace("nodejs", "node js")
    value = value.replace("postgresql", "postgre sql")
    value = value.replace("snowflakes", "snowflake")
    value = value.replace("powerplatform", "power platform")
    value = re.sub(r"[^a-z0-9+#.\s]", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def is_likely_company_location(candidate_info: dict) -> bool:
    source = normalize_text(candidate_info.get("candidate_location_source"))

    company_location_sources = [
        "work experience",
        "company",
        "employer",
        "project",
        "education",
        "college",
        "client",
        "office",
        "job",
        "not mentioned",
    ]

    return source in company_location_sources


def split_skills(value) -> List[str]:
    value = clean_cell(value)
    parts = re.split(r"[,~|;/\n]+", value)
    return [p.strip() for p in parts if p.strip()]


def safe_number(value, default=0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default
    
    
    
def select_relevant_matches_or_best_one(matches: list) -> list:
    """
    Business rule for Option 2:
    - If candidate fits multiple jobs, return all relevant matches.
    - If candidate fits nothing, return only the closest/best one.
    - Do not show all Not Suitable rows.
    """

    if not matches:
        return []

    cleaned_matches = []

    for match in matches:
        if not isinstance(match, dict):
            continue

        ats = safe_number(match.get("ats_score"))
        verdict = clean_cell(match.get("verdict"))
        final_remark = clean_cell(match.get("final_remark"))

        match["ats_score"] = ats
        match["verdict"] = verdict or "Not Suitable"
        match["final_remark"] = final_remark or "Not Suitable"
        match["call_status"] = ""

        cleaned_matches.append(match)

    if not cleaned_matches:
        return []

    # Sort highest ATS first
    cleaned_matches = sorted(
        cleaned_matches,
        key=lambda x: safe_number(x.get("ats_score")),
        reverse=True,
    )

    relevant_matches = []

    for match in cleaned_matches:
        ats = safe_number(match.get("ats_score"))
        verdict = clean_cell(match.get("verdict"))
        final_remark = clean_cell(match.get("final_remark"))

        is_relevant = (
            ats >= 50
            and verdict != "Not Suitable"
            and final_remark != "Not Suitable"
        )

        if is_relevant:
            relevant_matches.append(match)

    # If candidate fits one or more jobs, show all relevant jobs only
    if relevant_matches:
        return relevant_matches

    # If candidate fits nothing, show only the closest/best one
    return [cleaned_matches[0]]


def clamp_int(value, min_value=0, max_value=100) -> int:
    try:
        value = int(round(float(value)))
    except Exception:
        value = 0

    return max(min_value, min(max_value, value))


def calculate_ats_from_breakdown(result: dict) -> int:
    """
    Generic ATS calculation.
    No hardcoded technology names.
    Uses Claude's component-level recruiter evaluation.
    """

    breakdown = result.get("ats_breakdown") or {}

    if not isinstance(breakdown, dict):
        return clamp_int(result.get("ats_score"), 0, 100)

    primary = clamp_int(breakdown.get("primary_core_skill_fit"), 0, 35)
    secondary = clamp_int(breakdown.get("secondary_skill_fit"), 0, 15)
    experience = clamp_int(breakdown.get("experience_fit"), 0, 20)
    role = clamp_int(breakdown.get("role_domain_alignment"), 0, 15)
    location = clamp_int(breakdown.get("location_availability_fit"), 0, 10)
    ctc_notice = clamp_int(breakdown.get("ctc_notice_fit"), 0, 5)

    final_ats = primary + secondary + experience + role + location + ctc_notice

    return clamp_int(final_ats, 0, 100)


ALLOWED_RESUME_EXTENSIONS = [".pdf", ".docx", ".doc"]


def get_file_extension(filename: str) -> str:
    filename = filename or ""
    return os.path.splitext(filename.lower())[1]


def is_supported_resume_upload(file: UploadFile) -> bool:
    filename = file.filename or ""
    ext = get_file_extension(filename)
    return ext in ALLOWED_RESUME_EXTENSIONS



def convert_office_file_to_pdf_bytes(file_bytes: bytes, original_filename: str) -> bytes:
    """
    Converts DOC/DOCX to PDF using LibreOffice headless inside Linux/Docker.
    Returns converted PDF bytes.
    """

    original_filename = os.path.basename(original_filename or "resume.docx")
    ext = get_file_extension(original_filename)

    if ext not in [".docx", ".doc"]:
        raise ValueError(f"Unsupported office file extension for conversion: {ext}")

    with tempfile.TemporaryDirectory() as temp_dir:
        input_path = os.path.join(temp_dir, original_filename)

        with open(input_path, "wb") as f:
            f.write(file_bytes)

        command = [
            "libreoffice",
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--convert-to",
            "pdf",
            "--outdir",
            temp_dir,
            input_path,
        ]

        result = subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=120,
        )

        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice conversion failed. stdout={result.stdout}, stderr={result.stderr}"
            )

        expected_pdf_name = os.path.splitext(original_filename)[0] + ".pdf"
        expected_pdf_path = os.path.join(temp_dir, expected_pdf_name)

        if not os.path.exists(expected_pdf_path):
            pdf_files = [
                f for f in os.listdir(temp_dir)
                if f.lower().endswith(".pdf")
            ]

            if not pdf_files:
                raise RuntimeError(
                    f"PDF conversion failed. No PDF created. stdout={result.stdout}, stderr={result.stderr}"
                )

            expected_pdf_path = os.path.join(temp_dir, pdf_files[0])

        with open(expected_pdf_path, "rb") as f:
            return f.read()


def get_pdf_bytes_from_resume_upload(file_bytes: bytes, original_filename: str) -> bytes:
    """
    Normalizes resume upload into PDF bytes.
    PDF stays direct.
    DOC/DOCX gets converted to PDF.
    """

    ext = get_file_extension(original_filename)

    if ext == ".pdf":
        return file_bytes

    if ext in [".docx", ".doc"]:
        return convert_office_file_to_pdf_bytes(file_bytes, original_filename)

    raise ValueError(f"Unsupported resume file type: {ext}")


def clean_json_text(raw: str) -> dict:
    clean = raw.strip()

    if clean.startswith("```json"):
        clean = clean.removeprefix("```json").strip()

    if clean.startswith("```"):
        clean = clean.removeprefix("```").strip()

    if clean.endswith("```"):
        clean = clean.removesuffix("```").strip()

    start = clean.find("{")
    end = clean.rfind("}")

    if start != -1 and end != -1:
        clean = clean[start : end + 1]

    return json.loads(clean)


# =========================================================
# EXCEL LOADING
# =========================================================

def load_requirement_df_from_path(excel_path: str) -> pd.DataFrame:
    debug(f"Loading job-specific requirement Excel: {excel_path}")

    if not os.path.exists(excel_path):
        raise HTTPException(
            status_code=400,
            detail=f"Requirement Excel not found for this job: {excel_path}",
        )

    xls = pd.ExcelFile(excel_path)
    debug(f"Available sheets: {xls.sheet_names}")

    matched_sheets = [
        s for s in xls.sheet_names if "requirement" in str(s).strip().lower()
    ]

    if matched_sheets:
        sheet_name = matched_sheets[0]
    else:
        sheet_name = xls.sheet_names[0]

    debug(f"Using requirement sheet: {sheet_name}")

    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    rename_map = {
        "Monthly Company Pay Rate": "Rate Card",
        "Monthly Company Pay Rate ": "Rate Card",
        "Anually Company Pay Rate": "Yearly Rate",
        "Annually Company Pay Rate": "Yearly Rate",
        "Annual Company Pay Rate": "Yearly Rate",
        "Request ID": "Request-ID",
        "Request Id": "Request-ID",
        "Work Location City": "Work Location City",
        "Work Location CDF": "Work Location CDF",
    }

    df = df.rename(columns=rename_map)

    missing_cols = [
        col for col in REQUIRED_REQUIREMENT_COLUMNS if col not in df.columns
    ]

    if missing_cols:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Invalid Requirement Excel format.",
                "missing_columns": missing_cols,
                "expected_columns": REQUIRED_REQUIREMENT_COLUMNS,
            },
        )

    df["Request-ID"] = df["Request-ID"].astype(str).str.strip()

    df = df[df["Request-ID"].astype(str).str.strip() != ""]
    df = df.dropna(how="all")

    optional_cols = [
        "Work Location City",
        "System Enhancements Required",
        "Candidate Annual CTC",
    ]

    for col in optional_cols:
        if col not in df.columns:
            df[col] = ""

    if df.empty:
        raise HTTPException(
            status_code=400,
            detail="Requirement Excel has no valid rows with Request-ID.",
        )

    debug(f"Job requirement dataframe shape: {df.shape}")
    return df


def load_requirement_df(cfg: Settings) -> pd.DataFrame:
    excel_path = UPLOADED_REQUIREMENT_PATH

    debug(f"Loading uploaded requirement Excel: {excel_path}")

    if not os.path.exists(excel_path):
        raise HTTPException(
            status_code=400,
            detail="Requirement Excel not uploaded. Please upload Requirement Excel first.",
        )

    xls = pd.ExcelFile(excel_path)
    debug(f"Available sheets: {xls.sheet_names}")

    matched_sheets = [
        s for s in xls.sheet_names if "requirement" in str(s).strip().lower()
    ]

    if matched_sheets:
        sheet_name = matched_sheets[0]
    else:
        sheet_name = xls.sheet_names[0]

    debug(f"Using requirement sheet: {sheet_name}")

    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    rename_map = {
        "Monthly Company Pay Rate": "Rate Card",
        "Monthly Company Pay Rate ": "Rate Card",
        "Anually Company Pay Rate": "Yearly Rate",
        "Annually Company Pay Rate": "Yearly Rate",
        "Annual Company Pay Rate": "Yearly Rate",
        "Request ID": "Request-ID",
        "Request Id": "Request-ID",
        "Work Location City": "Work Location City",
        "Work Location CDF": "Work Location CDF",
    }

    df = df.rename(columns=rename_map)

    missing_cols = [
        col for col in REQUIRED_REQUIREMENT_COLUMNS if col not in df.columns
    ]

    if missing_cols:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Invalid Requirement Excel format.",
                "missing_columns": missing_cols,
                "expected_columns": REQUIRED_REQUIREMENT_COLUMNS,
            },
        )

    df["Request-ID"] = df["Request-ID"].astype(str).str.strip()

    df = df[df["Request-ID"].astype(str).str.strip() != ""]
    df = df.dropna(how="all")

    optional_cols = [
        "Work Location City",
        "System Enhancements Required",
        "Candidate Annual CTC",
    ]

    for col in optional_cols:
        if col not in df.columns:
            df[col] = ""

    if df.empty:
        raise HTTPException(
            status_code=400,
            detail="Requirement Excel has no valid rows with Request-ID.",
        )

    debug(f"Requirement dataframe shape: {df.shape}")
    debug(f"Requirement columns: {list(df.columns)}")

    return df


def get_requirement_by_id(request_id: str, cfg: Settings) -> dict:
    df = load_requirement_df(cfg)

    request_id = str(request_id).strip()
    matched = df[df["Request-ID"] == request_id]

    if matched.empty:
        raise HTTPException(
            status_code=404,
            detail=f"Request-ID not found in requirement Excel: {request_id}",
        )

    return matched.iloc[0].fillna("").to_dict()


# =========================================================
# CLAUDE CLIENT
# =========================================================


def get_claude_client(cfg: Settings) -> anthropic.Anthropic:
    return anthropic.Anthropic(api_key=cfg.anthropic_api_key)


# =========================================================
# CLAUDE - RESUME EXTRACTION
# =========================================================


def call_claude_extract_candidate(
    pdf_bytes: bytes,
    filename: str,
    cfg: Settings,
) -> dict:
    debug(f"Claude resume extraction started: {filename}")

    client = get_claude_client(cfg)

    pdf_base64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

    #     prompt = """
    # You are an expert resume parser for recruitment screening.

    # Read the uploaded resume PDF and extract candidate details.

    # Return only valid JSON. No markdown. No explanation.

    # Required JSON:
    # {
    #   "candidate_name": null,
    #   "candidate_phone": null,
    #   "candidate_email": null,
    #   "candidate_location": null,
    #   "candidate_total_experience_years": null,
    #   "candidate_current_ctc": null,
    #   "candidate_expected_ctc": null,
    #   "candidate_notice_period": null,
    #   "candidate_skills": [],
    #   "recent_job_title": null,
    #   "profile_summary": null
    # }

    # Rules:
    # - Do not hallucinate missing values.
    # - If value is missing, return null.
    # - candidate_total_experience_years should be a number if possible.
    # - candidate_skills should include programming languages, tools, frameworks, databases, cloud, platforms, domain skills and business tools.
    # - Extract from resume only.
    # """

    prompt = """
You are an expert resume parser for recruitment screening.

Read the uploaded resume PDF and extract candidate details accurately.

Return only valid JSON. No markdown. No explanation.

Required JSON:
{
  "candidate_name": null,
  "candidate_phone": null,
  "candidate_email": null,

  "candidate_location": null,
  "candidate_city": null,
  "candidate_state": null,
  "candidate_country": null,
  "candidate_location_source": null,
  "candidate_location_evidence": null,
  "candidate_location_confidence": "not_found",

  "candidate_total_experience_years": null,
  "candidate_current_ctc": null,
  "candidate_expected_ctc": null,
  "candidate_notice_period": null,

  "candidate_skills": [],
  "recent_job_title": null,
  "profile_summary": null
}

Strict extraction rules:
- Extract values only from the resume.
- Do not hallucinate missing values.
- If value is missing, return null.
- candidate_total_experience_years must be a number if possible.

Location extraction rules:
- candidate_location means candidate's own current/residential/preferred/contact/header/address location only.
- Extract candidate location only if it is clearly mentioned in the resume header, contact section, address section, personal details section, or explicitly written as Current Location / Location / Address / Residence / Preferred Location.
- Do NOT use company location, employer location, project location, client location, office location, college location, education location, or job experience location as candidate location.
- Do NOT infer candidate location from latest employer location.
- Do NOT infer candidate location from repeated work locations.
- If only work experience/company/project/education locations are mentioned, return candidate_location = null and candidate_city = null.
- candidate_location_evidence must contain the exact short text from the resume that proves candidate location.
- If there is no explicit candidate location evidence, return candidate_location_evidence = null.
- candidate_location_source must be one of:
  "header/contact"
  "address/personal details"
  "current location"
  "preferred location"
  "work experience"
  "education"
  "not mentioned"
- If the only location found comes from work experience or education, do not set candidate_location. Set candidate_location_source = "not mentioned".
- candidate_location_confidence must be one of:
  "high"
  "medium"
  "low"
  "not_found"
- Use "high" only when location is explicitly candidate's own location from header/contact/address/current location/preferred location.
- Use "medium" when location appears near contact details but section is not clearly labelled.
- Use "low" when location is ambiguous.
- Use "not_found" when candidate location is not explicitly mentioned.
- Do not put country/state inside candidate_city.
- If resume says "Mumbai, India" in candidate contact/address section:
  candidate_location = "Mumbai, India"
  candidate_city = "Mumbai"
  candidate_state = null
  candidate_country = "India"
  candidate_location_source = "header/contact"
  candidate_location_evidence = "Mumbai, India"
  candidate_location_confidence = "high"

Skill rules:
- candidate_skills should include programming languages, tools, frameworks, databases, cloud, platforms, domain skills and business tools.
- Keep candidate_skills clean and comma-independent.
- Do not include full sentences in candidate_skills.
"""

    message = client.messages.create(
        model=cfg.claude_model,
        max_tokens=cfg.max_tokens,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_base64,
                        },
                    },
                    {
                        "type": "text",
                        "text": prompt,
                    },
                ],
            }
        ],
    )

    raw = message.content[0].text.strip()
    usage = getattr(message, "usage", None)
    input_tokens = getattr(usage, "input_tokens", 0) if usage else 0
    output_tokens = getattr(usage, "output_tokens", 0) if usage else 0

    try:
        result = clean_json_text(raw)
    except Exception as e:
        debug(f"Candidate JSON parse failed for {filename}: {repr(e)}")
        # result = {
        #     "candidate_name": None,
        #     "candidate_phone": None,
        #     "candidate_email": None,
        #     "candidate_location": None,
        #     "candidate_total_experience_years": None,
        #     "candidate_current_ctc": None,
        #     "candidate_expected_ctc": None,
        #     "candidate_notice_period": None,
        #     "candidate_skills": [],
        #     "recent_job_title": None,
        #     "profile_summary": raw,
        # }

        result = {
            "candidate_name": None,
            "candidate_phone": None,
            "candidate_email": None,
            "candidate_location": None,
            "candidate_city": None,
            "candidate_state": None,
            "candidate_country": None,
            "candidate_location_source": None,
            "candidate_total_experience_years": None,
            "candidate_current_ctc": None,
            "candidate_expected_ctc": None,
            "candidate_notice_period": None,
            "candidate_skills": [],
            "recent_job_title": None,
            "profile_summary": raw,
        }

    if not isinstance(result.get("candidate_skills"), list):
        result["candidate_skills"] = []

    debug(f"Claude resume extraction completed: {filename}")

    result["_token_usage"] = {
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": input_tokens + output_tokens,
    }

    return result


# =========================================================
# PYTHON SHORTLISTING OF REQUIREMENTS
# =========================================================


def build_requirement_search_text(row: dict) -> str:
    return " ".join(
        [
            clean_cell(row.get("Job Title")),
            clean_cell(row.get("Skills - Name")),
            clean_cell(row.get("Skills - Experience")),
            clean_cell(row.get("Additional Skills")),
            clean_cell(row.get("Job Description")),
            clean_cell(row.get("Work Location CDF")),
        ]
    )


def build_candidate_search_text(candidate_info: dict) -> str:
    return " ".join(
        [
            " ".join(candidate_info.get("candidate_skills") or []),
            clean_cell(candidate_info.get("recent_job_title")),
            clean_cell(candidate_info.get("profile_summary")),
            clean_cell(candidate_info.get("candidate_location")),
        ]
    )


def shortlist_requirements(
    requirements_df: pd.DataFrame,
    candidate_info: dict,
    top_n: int = 25,
) -> list:
    candidate_text = build_candidate_search_text(candidate_info)
    candidate_norm = normalize_text(candidate_text)
    candidate_tokens = set(candidate_norm.split())

    scored_rows = []

    for _, row in requirements_df.iterrows():
        row_dict = row.fillna("").to_dict()

        req_text = build_requirement_search_text(row_dict)
        req_norm = normalize_text(req_text)
        req_tokens = set(req_norm.split())

        overlap = candidate_tokens.intersection(req_tokens)
        score = len(overlap)

        # Strong boost on exact primary skill tokens
        primary_skills = split_skills(row_dict.get("Skills - Name"))
        for skill in primary_skills:
            skill_norm = normalize_text(skill)
            if skill_norm and skill_norm in candidate_norm:
                score += 25

        # Boost additional skills
        additional_skills = split_skills(row_dict.get("Additional Skills"))
        for skill in additional_skills[:80]:
            skill_norm = normalize_text(skill)
            if skill_norm and skill_norm in candidate_norm:
                score += 5

        # Prefer open requirements slightly
        status = normalize_text(row_dict.get("Status"))
        if status == "open":
            score += 5

        scored_rows.append((score, row_dict))

    scored_rows = sorted(scored_rows, key=lambda x: x[0], reverse=True)

    top_rows = [row for score, row in scored_rows[:top_n]]

    debug(f"Shortlisted {len(top_rows)} requirements")
    return top_rows


# =========================================================
# CLAUDE - BEST REQUIREMENT MATCHING
# =========================================================


def requirement_for_prompt(row: dict) -> dict:
    return {
        "Request-ID": clean_cell(row.get("Request-ID")),
        "MSP Owner": clean_cell(row.get("MSP Owner")),
        "Job Title": clean_cell(row.get("Job Title")),
        "Skills - Name": clean_cell(row.get("Skills - Name")),
        "Skills - Experience": clean_cell(row.get("Skills - Experience")),
        "Additional Skills": clean_cell(row.get("Additional Skills")),
        "Job Description": clean_cell(row.get("Job Description")),
        "Status": clean_cell(row.get("Status")),
        "Work Location City": clean_cell(row.get("Work Location City")),
        "Work Location CDF": clean_cell(row.get("Work Location CDF")),
        "Rate Card": clean_cell(row.get("Rate Card")),
        "Yearly Rate": clean_cell(row.get("Yearly Rate")),
        "System Enhancements Required": clean_cell(
            row.get("System Enhancements Required")
        ),
        "Candidate Annual CTC": clean_cell(row.get("Candidate Annual CTC")),
    }


# def call_claude_best_requirement_match(
#     candidate_info: dict,
#     shortlisted_requirements: list,
#     cfg: Settings,
# ) -> dict:
#     debug("Claude best requirement matching started")

#     client = get_claude_client(cfg)

#     requirements_for_prompt = [
#         requirement_for_prompt(row) for row in shortlisted_requirements
#     ]

#     prompt = f"""
# You are an expert technical recruiter.

# You are given:
# 1. One candidate profile extracted from resume.
# 2. A shortlisted list of job requirements from Requirement Sheet.

# Your task:
# - Select the best matching Request-ID/job requirement for the candidate.
# - If no requirement is a good fit, select the closest Request-ID but mark final_remark as "Not Suitable".
# - Generate ATS score from 0 to 100.
# - Generate recruiter-style verdict, call_status and remark.

# Important:
# - Use ONLY the provided requirements.
# - Do not invent Request-ID.
# - NPU means Not Picked Up. It cannot be detected from resume. Do not set NPU unless input clearly says call status is NPU.
# - Use High CTC only if candidate CTC is available and exceeds requirement budget.
# - Use High Notice Period only if notice period is available and clearly high.
# - Use Location Mismatch if location is main issue.
# - Use Skill mismatch if skill gap is main issue.
# - Use Exp mismatch if experience is main issue.
# - Use Match only when candidate is a strong/clear fit.
# - Use Not Suitable when profile/domain is not aligned.
# - call_status must always be empty string "" because call status is recruiter-entered manually after calling the candidate.


# ATS scoring rules:
# - ATS must be calculated using the scoring breakdown below.
# - Do not give a random or generic ATS score.
# - Do not give the same ATS score to different candidates unless their breakdown is genuinely the same.
# - Do not hardcode technology names. Score only based on the selected requirement text and candidate profile.

# ATS breakdown:
# 1. primary_core_skill_fit: 0 to 35
#    - 30-35: Candidate strongly matches the role-defining/core skills.
#    - 20-29: Candidate matches most core skills but has some gaps.
#    - 10-19: Candidate has partial core skill overlap.
#    - 0-9: Candidate does not match the main/core skill need.

# 2. secondary_skill_fit: 0 to 15
#    - 12-15: Strong supporting skill match.
#    - 7-11: Partial supporting skill match.
#    - 0-6: Weak or missing supporting skills.

# 3. experience_fit: 0 to 20
#    - 18-20: Candidate is within required experience range.
#    - 12-17: Candidate is slightly below/above required range.
#    - 0-11: Major experience mismatch.

# 4. role_domain_alignment: 0 to 15
#    - 12-15: Candidate's recent role/domain is very aligned.
#    - 7-11: Candidate is related but not exact.
#    - 0-6: Candidate profile/domain is not aligned.

# 5. location_availability_fit: 0 to 10
#    - 8-10: Location/preference clearly matches.
#    - 5-7: Location is not evaluated or flexible/unclear.
#    - 0-4: Clear location mismatch.

# 6. ctc_notice_fit: 0 to 5
#    - 5: CTC and notice are suitable or not available.
#    - 3-4: Minor concern.
#    - 0-2: Clear CTC or notice period issue.

# Final ATS score must equal:
# primary_core_skill_fit + secondary_skill_fit + experience_fit + role_domain_alignment + location_availability_fit + ctc_notice_fit

# Verdict should follow final ATS:
# - 85-100 = Strong Fit
# - 70-84 = Good Fit
# - 50-69 = Possible Fit
# - 0-49 = Not Suitable


# Allowed verdict values:
# - Strong Fit
# - Good Fit
# - Possible Fit
# - Not Suitable

# Allowed final_remark values:
# - Exp mismatch
# - High CTC
# - High Notice Period
# - Location Mismatch
# - Match
# - Not Suitable
# - Skill mismatch

# Set experience_mismatch as "Yes" if candidate total experience is less than required experience, otherwise "No".
# Set skill_mismatch as "Yes" if candidate is missing important required skills, otherwise "No".

# Location rules:
# - Compare candidate location with requirement location only when candidate_location_confidence is "high" or "medium".
# - Do not compare location if candidate_location_confidence is "low" or "not_found".
# - If candidate_location is null/blank/not_found, set location_mismatch = "Not Evaluated".
# - If location is not evaluated, do not use final_remark = "Location Mismatch".
# - Compare locations semantically and case-insensitively.
# - Ignore formatting differences, symbols, punctuation, country/state suffixes, office names, SEZ/branch/building names, and extra whitespace.
# - Treat equivalent city spellings or common regional names as same when clearly equivalent.
# - Do not infer candidate location from employer/work/education/project location.
# - In reason, if candidate location is missing, write:
#   "Candidate location is not explicitly mentioned in the resume, so location fit is not evaluated."
  
  


# Candidate Profile:
# {json.dumps(candidate_info, ensure_ascii=False, indent=2)}

# Shortlisted Requirements:
# {json.dumps(requirements_for_prompt, ensure_ascii=False, indent=2)}

# Return only valid JSON:
# {{
#   "best_request_id": null,
#   "ats_score": 0,
#   "ats_breakdown": {{
#     "primary_core_skill_fit": 0,
#     "secondary_skill_fit": 0,
#     "experience_fit": 0,
#     "role_domain_alignment": 0,
#     "location_availability_fit": 0,
#     "ctc_notice_fit": 0
#   }},
#   "verdict": "Not Suitable",
#   "call_status": "",
#   "final_remark": "Not Suitable",
#   "experience_mismatch": "No",
#   "skill_mismatch": "No",
#   "location_mismatch": "Not Evaluated",
#   "matching_skills": [],
#   "missing_skills": [],
#   "reason": ""
# }}
# """

#     message = client.messages.create(
#         model=cfg.claude_model,
#         max_tokens=cfg.max_tokens,
#         messages=[
#             {
#                 "role": "user",
#                 "content": [
#                     {
#                         "type": "text",
#                         "text": prompt,
#                     }
#                 ],
#             }
#         ],
#     )

#     raw = message.content[0].text.strip()

#     usage = getattr(message, "usage", None)
#     input_tokens = getattr(usage, "input_tokens", 0) if usage else 0
#     output_tokens = getattr(usage, "output_tokens", 0) if usage else 0

#     try:
#         result = clean_json_text(raw)
#     except Exception as e:
#         debug(f"Match JSON parse failed: {repr(e)}")
#         # result = {
#         #     "best_request_id": None,
#         #     "ats_score": 0,
#         #     "verdict": "Not Suitable",
#         #     "call_status": "",
#         #     "final_remark": "Not Suitable",
#         #     "matching_skills": [],
#         #     "missing_skills": [],
#         #     "reason": raw,
#         # }
#         result = {
#             "best_request_id": None,
#             "ats_score": 0,
#             "ats_breakdown": {
#                 "primary_core_skill_fit": 0,
#                 "secondary_skill_fit": 0,
#                 "experience_fit": 0,
#                 "role_domain_alignment": 0,
#                 "location_availability_fit": 0,
#                 "ctc_notice_fit": 0,
#             },
#             "verdict": "Not Suitable",
#             "call_status": "",
#             "final_remark": "Not Suitable",
#             "experience_mismatch": "No",
#             "skill_mismatch": "No",
#             "location_mismatch": "Not Evaluated",
#             "matching_skills": [],
#             "missing_skills": [],
#             "reason": raw,
#         }

#     final_remark = clean_cell(result.get("final_remark"))

#     if final_remark not in VALID_REMARKS:
#         final_remark = "Not Suitable"

#     result["final_remark"] = final_remark
#     # Calculate ATS from component breakdown so score is not random.
#     result["ats_score"] = calculate_ats_from_breakdown(result)

#     # verdict = clean_cell(result.get("verdict"))
#     # if verdict not in ["Strong Fit", "Good Fit", "Possible Fit", "Not Suitable"]:
#     #     ats = safe_number(result.get("ats_score"))
#     #     if ats >= 80:
#     #         verdict = "Strong Fit"
#     #     elif ats >= 60:
#     #         verdict = "Good Fit"
#     #     elif ats >= 40:
#     #         verdict = "Possible Fit"
#     #     else:
#     #         verdict = "Not Suitable"

#     # result["verdict"] = verdict
    
#     ats = safe_number(result.get("ats_score"))

#     if ats >= 85:
#         verdict = "Strong Fit"
#     elif ats >= 70:
#         verdict = "Good Fit"
#     elif ats >= 50:
#         verdict = "Possible Fit"
#     else:
#         verdict = "Not Suitable"

#     result["verdict"] = verdict

#     debug("Claude best requirement matching completed")

#     result["_token_usage"] = {
#         "input_tokens": input_tokens,
#         "output_tokens": output_tokens,
#         "total_tokens": input_tokens + output_tokens,
#     }

#     return result



def call_claude_best_requirement_match(
    candidate_info: dict,
    shortlisted_requirements: list,
    cfg: Settings,
) -> dict:
    debug("Claude best requirement matching started")

    client = get_claude_client(cfg)

    requirements_for_prompt = [
        requirement_for_prompt(row) for row in shortlisted_requirements
    ]

    prompt = f"""
You are an expert technical recruiter.

You are given:
1. One candidate profile extracted from resume.
2. A shortlisted list of job requirements from Requirement Sheet.

Your task:
- Select the best matching Request-ID/job requirement for the candidate.
- If no requirement is a good fit, select the closest Request-ID but mark final_remark as "Not Suitable".
- Generate ATS score from 0 to 100 based on overall recruiter judgment.
- Generate recruiter-style verdict, call_status and remark.

Important:
- Use ONLY the provided requirements.
- Do not invent Request-ID.
- NPU means Not Picked Up. It cannot be detected from resume. Do not set NPU unless input clearly says call status is NPU.
- Use High CTC only if candidate CTC is available and exceeds requirement budget.
- Use High Notice Period only if notice period is available and clearly high.
- Use Location Mismatch if location is main issue.
- Use Skill mismatch if skill gap is main issue.
- Use Exp mismatch if experience is main issue.
- Use Match only when candidate is a strong/clear fit.
- Use Not Suitable when profile/domain is not aligned.
- call_status must always be empty string "" because call status is recruiter-entered manually after calling the candidate.

ATS scoring guidance:
- ATS must reflect overall match quality between the candidate and selected requirement.
- Do not give the same ATS score to every candidate unless their fit is genuinely almost identical.
- Give lower score when profile/domain is weakly aligned.
- Give medium score when candidate is partially aligned but has some important gaps.
- Give higher score when candidate strongly matches experience, core skills, role focus, and location/availability where available.
- Keep ATS realistic and recruiter-friendly.

Location rules:
- Compare candidate location with requirement location only when candidate_location_confidence is "high" or "medium".
- Do not compare location if candidate_location_confidence is "low" or "not_found".
- If candidate_location is null/blank/not_found, set location_mismatch = "Not Evaluated".
- If location is not evaluated, do not use final_remark = "Location Mismatch".
- Compare locations semantically and case-insensitively.
- Ignore formatting differences, symbols, punctuation, country/state suffixes, office names, SEZ/branch/building names, and extra whitespace.
- Treat equivalent city spellings or common regional names as same when clearly equivalent.
- Do not infer candidate location from employer/work/education/project location.
- In reason, if candidate location is missing, write:
  "Candidate location is not explicitly mentioned in the resume, so location fit is not evaluated."

Allowed verdict values:
- Strong Fit
- Good Fit
- Possible Fit
- Not Suitable

Allowed final_remark values:
- Exp mismatch
- High CTC
- High Notice Period
- Location Mismatch
- Match
- Not Suitable
- Skill mismatch

Set experience_mismatch as "Yes" if candidate total experience is less than required experience, otherwise "No".
Set skill_mismatch as "Yes" if candidate is missing important required skills, otherwise "No".
Set location_mismatch as "Yes" only if candidate location is explicitly available and conflicts with requirement location.
If candidate location is missing, set location_mismatch as "Not Evaluated".

Candidate Profile:
{json.dumps(candidate_info, ensure_ascii=False, indent=2)}

Shortlisted Requirements:
{json.dumps(requirements_for_prompt, ensure_ascii=False, indent=2)}

Return only valid JSON:
{{
  "best_request_id": null,
  "ats_score": 0,
  "verdict": "Not Suitable",
  "call_status": "",
  "final_remark": "Not Suitable",
  "experience_mismatch": "No",
  "skill_mismatch": "No",
  "location_mismatch": "Not Evaluated",
  "matching_skills": [],
  "missing_skills": [],
  "reason": ""
}}
"""

    message = client.messages.create(
        model=cfg.claude_model,
        max_tokens=cfg.max_tokens,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt,
                    }
                ],
            }
        ],
    )

    raw = message.content[0].text.strip()

    usage = getattr(message, "usage", None)
    input_tokens = getattr(usage, "input_tokens", 0) if usage else 0
    output_tokens = getattr(usage, "output_tokens", 0) if usage else 0

    try:
        result = clean_json_text(raw)
    except Exception as e:
        debug(f"Match JSON parse failed: {repr(e)}")
        result = {
            "best_request_id": None,
            "ats_score": 0,
            "verdict": "Not Suitable",
            "call_status": "",
            "final_remark": "Not Suitable",
            "experience_mismatch": "No",
            "skill_mismatch": "No",
            "location_mismatch": "Not Evaluated",
            "matching_skills": [],
            "missing_skills": [],
            "reason": raw,
        }

    # Always keep call status blank. Recruiter will fill this manually.
    result["call_status"] = ""

    final_remark = clean_cell(result.get("final_remark"))

    if final_remark not in VALID_REMARKS:
        final_remark = "Not Suitable"

    result["final_remark"] = final_remark

    ats = safe_number(result.get("ats_score"))

    # Keep verdict aligned with ATS, but do not force ATS itself.
    if ats >= 85:
        verdict = "Strong Fit"
    elif ats >= 70:
        verdict = "Good Fit"
    elif ats >= 50:
        verdict = "Possible Fit"
    else:
        verdict = "Not Suitable"

    result["verdict"] = verdict

    # Safety: if candidate location is missing, never allow Location Mismatch.
    candidate_city = clean_cell(candidate_info.get("candidate_city"))
    candidate_location = clean_cell(candidate_info.get("candidate_location"))

    if not candidate_city and not candidate_location:
        if result.get("final_remark") == "Location Mismatch":
            result["final_remark"] = "Not Suitable"

        result["location_mismatch"] = "Not Evaluated"

        reason = clean_cell(result.get("reason"))
        if "location fit is not evaluated" not in reason.lower():
            result["reason"] = (
                reason
                + " Candidate location is not explicitly mentioned in the resume, so location fit is not evaluated."
            ).strip()

    debug("Claude best requirement matching completed")

    result["_token_usage"] = {
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": input_tokens + output_tokens,
    }

    return result



def call_claude_relevant_requirement_matches(
    candidate_info: dict,
    shortlisted_requirements: list,
    cfg: Settings,
) -> dict:
    debug("Claude relevant requirement matching started")

    client = get_claude_client(cfg)

    requirements_for_prompt = [
        requirement_for_prompt(row) for row in shortlisted_requirements
    ]

    prompt = f"""
You are an expert technical recruiter.

You are given:
1. One candidate profile extracted from resume.
2. A shortlisted list of job requirements from Requirement Sheet.

Your task:
- Evaluate the candidate against every shortlisted Request-ID/job requirement.
- Return all relevant matching Request-IDs/job requirements for the candidate.
- If no requirement is a good fit, return the closest Request-ID but mark final_remark as "Not Suitable".
- Generate ATS score from 0 to 100 based on overall recruiter judgment for each returned requirement.
- Generate recruiter-style verdict, call_status and remark for each returned requirement.

Important:
- Use ONLY the provided requirements.
- Do not invent Request-ID.
- A candidate can be mapped to multiple requirements when relevant.
- Return one match object per relevant Request-ID.
- NPU means Not Picked Up. It cannot be detected from resume. Do not set NPU unless input clearly says call status is NPU.
- Use High CTC only if candidate CTC is available and exceeds requirement budget.
- Use High Notice Period only if notice period is available and clearly high.
- Use Location Mismatch if location is main issue.
- Use Skill mismatch if skill gap is main issue.
- Use Exp mismatch if experience is main issue.
- Use Match only when candidate is a strong/clear fit.
- Use Not Suitable when profile/domain is not aligned.
- call_status must always be empty string "" because call status is recruiter-entered manually after calling the candidate.

Relevance rules:
- Return all requirements that are relevant enough for recruiter review.
- Relevant means the candidate has meaningful role, skill, experience, or domain overlap with the requirement.
- Normally include matches with verdict "Possible Fit", "Good Fit", or "Strong Fit".
- Do not return weak unrelated requirements just to increase rows.
- If no requirement is relevant, return exactly one closest requirement as "Not Suitable" for tracking.

ATS scoring guidance:
- ATS must reflect overall match quality between the candidate and each selected requirement.
- Do not give the same ATS score to every requirement unless their fit is genuinely almost identical.
- Give lower score when profile/domain is weakly aligned.
- Give medium score when candidate is partially aligned but has some important gaps.
- Give higher score when candidate strongly matches experience, core skills, role focus, and location/availability where available.
- Keep ATS realistic and recruiter-friendly.

Location rules:
- Compare candidate location with requirement location only when candidate_location_confidence is "high" or "medium".
- Do not compare location if candidate_location_confidence is "low" or "not_found".
- If candidate_location is null/blank/not_found, set location_mismatch = "Not Evaluated".
- If location is not evaluated, do not use final_remark = "Location Mismatch".
- Compare locations semantically and case-insensitively.
- Ignore formatting differences, symbols, punctuation, country/state suffixes, office names, SEZ/branch/building names, and extra whitespace.
- Treat equivalent city spellings or common regional names as same when clearly equivalent.
- Do not infer candidate location from employer/work/education/project location.
- In reason, if candidate location is missing, write:
  "Candidate location is not explicitly mentioned in the resume, so location fit is not evaluated."

Allowed verdict values:
- Strong Fit
- Good Fit
- Possible Fit
- Not Suitable

Allowed final_remark values:
- Exp mismatch
- High CTC
- High Notice Period
- Location Mismatch
- Match
- Not Suitable
- Skill mismatch

Set experience_mismatch as "Yes" if candidate total experience is less than required experience, otherwise "No".
Set skill_mismatch as "Yes" if candidate is missing important required skills, otherwise "No".
Set location_mismatch as "Yes" only if candidate location is explicitly available and conflicts with requirement location.
If candidate location is missing, set location_mismatch as "Not Evaluated".

Candidate Profile:
{json.dumps(candidate_info, ensure_ascii=False, indent=2)}

Shortlisted Requirements:
{json.dumps(requirements_for_prompt, ensure_ascii=False, indent=2)}

Return only valid JSON:
{{
  "matches": [
    {{
      "request_id": null,
      "ats_score": 0,
      "verdict": "Not Suitable",
      "call_status": "",
      "final_remark": "Not Suitable",
      "experience_mismatch": "No",
      "skill_mismatch": "No",
      "location_mismatch": "Not Evaluated",
      "matching_skills": [],
      "missing_skills": [],
      "reason": ""
    }}
  ]
}}
"""

    message = client.messages.create(
        model=cfg.claude_model,
        max_tokens=cfg.max_tokens,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt,
                    }
                ],
            }
        ],
    )

    raw = message.content[0].text.strip()

    usage = getattr(message, "usage", None)
    input_tokens = getattr(usage, "input_tokens", 0) if usage else 0
    output_tokens = getattr(usage, "output_tokens", 0) if usage else 0

    try:
        result = clean_json_text(raw)
    except Exception as e:
        debug(f"Multi match JSON parse failed: {repr(e)}")
        result = {"matches": []}

    matches = result.get("matches")
    if not isinstance(matches, list):
        matches = []

    cleaned_matches = []

    for match in matches:
        if not isinstance(match, dict):
            continue

        # Always keep call status blank. Recruiter will fill this manually.
        match["call_status"] = ""

        request_id = clean_cell(match.get("request_id")) or clean_cell(
            match.get("best_request_id")
        )
        match["request_id"] = request_id
        match["best_request_id"] = request_id

        final_remark = clean_cell(match.get("final_remark"))
        if final_remark not in VALID_REMARKS:
            final_remark = "Not Suitable"
        match["final_remark"] = final_remark

        ats = safe_number(match.get("ats_score"))

        # Keep verdict aligned with ATS, but do not force ATS itself.
        if ats >= 85:
            verdict = "Strong Fit"
        elif ats >= 70:
            verdict = "Good Fit"
        elif ats >= 50:
            verdict = "Possible Fit"
        else:
            verdict = "Not Suitable"

        match["verdict"] = verdict

        # Safety: if candidate location is missing, never allow Location Mismatch.
        candidate_city = clean_cell(candidate_info.get("candidate_city"))
        candidate_location = clean_cell(candidate_info.get("candidate_location"))

        if not candidate_city and not candidate_location:
            if match.get("final_remark") == "Location Mismatch":
                match["final_remark"] = "Not Suitable"

            match["location_mismatch"] = "Not Evaluated"

            reason = clean_cell(match.get("reason"))
            if "location fit is not evaluated" not in reason.lower():
                match["reason"] = (
                    reason
                    + " Candidate location is not explicitly mentioned in the resume, so location fit is not evaluated."
                ).strip()

        cleaned_matches.append(match)

    result["matches"] = cleaned_matches

    debug("Claude relevant requirement matching completed")

    result["_token_usage"] = {
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": input_tokens + output_tokens,
    }

    return result

# =========================================================
# OUTPUT ROW BUILDERS
# =========================================================


def find_requirement_by_request_id(
    requirements_df: pd.DataFrame, request_id: str
) -> dict:
    request_id = clean_cell(request_id)

    if not request_id:
        return {}

    matched = requirements_df[
        requirements_df["Request-ID"].astype(str).str.strip() == request_id
    ]

    if matched.empty:
        return {}

    return matched.iloc[0].fillna("").to_dict()


def build_output_row(
    req: dict, candidate_info: dict, match_result: dict, filename: str = ""
) -> dict:
    candidate_skills = candidate_info.get("candidate_skills") or []

    candidate_city = clean_cell(candidate_info.get("candidate_city"))
    candidate_location = clean_cell(candidate_info.get("candidate_location"))

    # Prefer clean city only if available.
    # If city is not explicitly found, keep location blank instead of using company location.
    final_candidate_location = candidate_city or candidate_location

    return {
        "Request-ID": clean_cell(req.get("Request-ID")),
        "MSP Owner": clean_cell(req.get("MSP Owner")),
        "Job Title": clean_cell(req.get("Job Title")),
        "Skills - Name": clean_cell(req.get("Skills - Name")),
        "Skills - Experience": clean_cell(req.get("Skills - Experience")),
        "Additional Skills": clean_cell(req.get("Additional Skills")),
        "Job Description": clean_cell(req.get("Job Description")),
        "Work Location CDF": clean_cell(req.get("Work Location CDF")),
        "Rate Card": clean_cell(req.get("Rate Card")),
        "Annually": clean_cell(req.get("Yearly Rate")),
        "CV File Name": clean_cell(filename),
        "Candidate Name": clean_cell(candidate_info.get("candidate_name")),
        "Candidate Phone": clean_cell(candidate_info.get("candidate_phone")),
        "Candidate Email": clean_cell(candidate_info.get("candidate_email")),
        "Candidate Location": final_candidate_location,
        "Candidate Total Experience": clean_cell(
            candidate_info.get("candidate_total_experience_years")
        ),
        "Candidate Skills": ", ".join(candidate_skills),
        "Experience Mismatch": clean_cell(match_result.get("experience_mismatch"))
        or "No",
        "Skill Mismatch": clean_cell(match_result.get("skill_mismatch")) or "No",
        "ATS": clean_cell(match_result.get("ats_score")),
        "Remark": clean_cell(match_result.get("reason")),
    }


def build_tracker_row(
    req: dict, candidate_info: dict, match_result: dict, filename: str
) -> dict:
    candidate_name = clean_cell(candidate_info.get("candidate_name"))

    if not candidate_name:
        candidate_name = filename

    return {
        "Date": datetime.now().strftime("%d-%m-%Y"),
        "Request ID": clean_cell(req.get("Request-ID"))
        or clean_cell(match_result.get("best_request_id")),
        "Status": clean_cell(req.get("Status")),
        "Skills": clean_cell(req.get("Skills - Name")),
        "Candidate": candidate_name,
        "Verdict": clean_cell(match_result.get("verdict")),
        "Call Status": "",
        "Remarks": clean_cell(match_result.get("final_remark")),
    }


# =========================================================
# EXCEL GENERATION
# =========================================================


def style_output_sheet(ws):
    header_blue = PatternFill("solid", fgColor="B7DEE8")
    candidate_green = PatternFill("solid", fgColor="DAF2D0")
    result_pink = PatternFill("solid", fgColor="F2CEEF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border

        if cell.column <= 10:
            cell.fill = header_blue
        elif 11 <= cell.column <= 16:
            cell.fill = candidate_green
        else:
            cell.fill = result_pink

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_tracker_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    verdict_col = None
    for cell in ws[1]:
        if cell.value == "Verdict":
            verdict_col = cell.column
            break

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

        if verdict_col:
            verdict_cell = row[verdict_col - 1]
            verdict = clean_cell(verdict_cell.value)
            color = VERDICT_COLORS.get(verdict)

            if color:
                verdict_cell.fill = PatternFill("solid", fgColor=color)
                verdict_cell.font = Font(bold=True, color="FFFFFF")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_pivot_sheet(ws):
    header_fill = PatternFill("solid", fgColor="B7DEE8")
    total_fill = PatternFill("solid", fgColor="B7DEE8")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

        first_cell = row[0]
        if clean_cell(first_cell.value) == "Grand Total":
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = total_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_adjust_columns(ws):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0

        for cell in col_cells:
            value = clean_cell(cell.value)
            if len(value) > max_len:
                max_len = len(value[:100])

        width = min(max(max_len + 2, 12), 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_pivot_df(tracker_df: pd.DataFrame) -> pd.DataFrame:
    counts = tracker_df["Remarks"].fillna("Not Suitable").value_counts().to_dict()

    rows = []

    for remark in PIVOT_REMARK_ORDER:
        count = int(counts.get(remark, 0))
        if count > 0:
            rows.append(
                {
                    "Row Labels": remark,
                    "Count of Remarks": count,
                }
            )

    grand_total = sum(row["Count of Remarks"] for row in rows)

    rows.append(
        {
            "Row Labels": "Grand Total",
            "Count of Remarks": grand_total,
        }
    )

    return pd.DataFrame(rows, columns=["Row Labels", "Count of Remarks"])


def create_final_excel(output_rows: list, tracker_rows: list, output_path: str):
    output_df = pd.DataFrame(output_rows, columns=OUTPUT_COLUMNS)
    tracker_df = pd.DataFrame(tracker_rows, columns=TRACKER_COLUMNS)
    pivot_df = create_pivot_df(tracker_df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        output_df.to_excel(writer, sheet_name="Output Sheet", index=False)
        tracker_df.to_excel(writer, sheet_name="Tracker", index=False)
        pivot_df.to_excel(writer, sheet_name="Pivot", index=False)

        workbook = writer.book

        ws_output = workbook["Output Sheet"]
        ws_tracker = workbook["Tracker"]
        ws_pivot = workbook["Pivot"]

        style_output_sheet(ws_output)
        style_tracker_sheet(ws_tracker)
        style_pivot_sheet(ws_pivot)

        auto_adjust_columns(ws_output)
        auto_adjust_columns(ws_tracker)
        auto_adjust_columns(ws_pivot)

    debug(f"Final Excel created: {output_path}")


def get_default_status():
    return {
        "job_id": "",
        "status": "idle",  # idle, queued, processing, completed, failed
        "message": "No active job",
        "total": 0,
        "processed": 0,
        "successful": 0,
        "failed": 0,
        "skipped": 0,
        "current_file": "",
        "current_batch": "",
        "output_filename": "",
        "download_url": "",
        "failed_files": [],
        "skipped_files": [],
        "token_usage": {
            "input_tokens": 0,
            "output_tokens": 0,
            "total_tokens": 0,
        },
        "started_at": "",
        "updated_at": "",
        "completed_at": "",
        "cost_info": {
            "model_name": "",
            "input_cost_usd": 0,
            "output_cost_usd": 0,
            "total_cost_usd": 0,
            "total_cost_inr": 0,
            "cost_per_resume_usd": 0,
            "cost_per_resume_inr": 0,
        },
        "runtime_info": {
            "started_at": "",
            "completed_at": "",
            "total_seconds": 0,
            "total_time_text": "",
            "average_seconds_per_resume": 0,
        },
        "resume_logs": [],
    }


def read_status():
    if not os.path.exists(ACTIVE_STATUS_PATH):
        return get_default_status()

    try:
        with open(ACTIVE_STATUS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return get_default_status()


def write_status(data: dict):
    data["updated_at"] = datetime.now().isoformat()

    with open(ACTIVE_STATUS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def is_active_job_running():
    status = read_status()
    return status.get("status") in ["queued", "processing"]


def reset_bulk_job_storage():
    if os.path.exists(ACTIVE_STATUS_PATH):
        os.remove(ACTIVE_STATUS_PATH)

    if os.path.exists(ACTIVE_UPLOAD_DIR):
        shutil.rmtree(ACTIVE_UPLOAD_DIR)

    os.makedirs(ACTIVE_UPLOAD_DIR, exist_ok=True)


class SavedUploadFile:
    def __init__(self, filename: str, content_type: str, path: str):
        self.filename = filename
        self.content_type = content_type
        self.path = path

    async def read(self):
        with open(self.path, "rb") as f:
            return f.read()


def calculate_cost_info(
    input_tokens: int,
    output_tokens: int,
    processed_count: int,
    model_name: str,
):
    # Keep these configurable later.
    # For now assuming Claude Sonnet pricing.
    input_rate_per_million = 3.0
    output_rate_per_million = 15.0

    # Approx USD-INR. You can update from config later.
    usd_to_inr = 95.4

    input_cost_usd = (input_tokens / 1_000_000) * input_rate_per_million
    output_cost_usd = (output_tokens / 1_000_000) * output_rate_per_million
    total_cost_usd = input_cost_usd + output_cost_usd
    total_cost_inr = total_cost_usd * usd_to_inr

    cost_per_resume_inr = 0
    cost_per_resume_usd = 0

    if processed_count > 0:
        cost_per_resume_inr = total_cost_inr / processed_count
        cost_per_resume_usd = total_cost_usd / processed_count

    return {
        "model_name": model_name,
        "input_rate_per_million_usd": input_rate_per_million,
        "output_rate_per_million_usd": output_rate_per_million,
        "usd_to_inr": usd_to_inr,
        "input_cost_usd": round(input_cost_usd, 4),
        "output_cost_usd": round(output_cost_usd, 4),
        "total_cost_usd": round(total_cost_usd, 4),
        "total_cost_inr": round(total_cost_inr, 2),
        "cost_per_resume_usd": round(cost_per_resume_usd, 4),
        "cost_per_resume_inr": round(cost_per_resume_inr, 2),
        "note": "Estimated cost based on configured pricing. Final provider billing may vary.",
    }


def format_duration(seconds: float):
    seconds = int(seconds)

    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60

    if hours > 0:
        return f"{hours} hr {minutes} min {secs} sec"

    if minutes > 0:
        return f"{minutes} min {secs} sec"

    return f"{secs} sec"


# =========================================================
# ROUTES
# =========================================================


@app.get("/")
def root():
    return {
        "message": "Resume Requirement Bulk Matching API",
        "usage": {
            "list_jobs": "GET /jobs",
            "get_job": "GET /jobs/{request_id}",
            "bulk_analyze": "POST /bulk-analyze with multiple PDF resume files; each resume maps to all relevant requirements",
            "download": "GET /download/{filename}",
        },
    }


@app.post("/upload-requirement")
async def upload_requirement_excel(
    file: UploadFile = File(...),
    cfg: Settings = Depends(get_settings),
):
    filename = file.filename or ""

    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Please upload a valid Excel file with .xlsx or .xls extension.",
        )

    content = await file.read()

    if not content:
        raise HTTPException(status_code=400, detail="Uploaded Excel file is empty.")

    temp_path = os.path.join(INPUT_DIR, f"temp_{uuid.uuid4().hex}_{filename}")

    try:
        with open(temp_path, "wb") as f:
            f.write(content)

        xls = pd.ExcelFile(temp_path)

        matched_sheets = [
            s for s in xls.sheet_names if "requirement" in str(s).strip().lower()
        ]

        if matched_sheets:
            sheet_name = matched_sheets[0]
        else:
            sheet_name = xls.sheet_names[0]

        df = pd.read_excel(temp_path, sheet_name=sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        rename_map = {
            "Monthly Company Pay Rate": "Rate Card",
            "Monthly Company Pay Rate ": "Rate Card",
            "Anually Company Pay Rate": "Yearly Rate",
            "Annually Company Pay Rate": "Yearly Rate",
            "Annual Company Pay Rate": "Yearly Rate",
            "Request ID": "Request-ID",
            "Request Id": "Request-ID",
            "Work Location City": "Work Location City",
            "Work Location CDF": "Work Location CDF",
        }

        df = df.rename(columns=rename_map)

        missing_cols = [
            col for col in REQUIRED_REQUIREMENT_COLUMNS if col not in df.columns
        ]

        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "Invalid Requirement Excel format.",
                    "missing_columns": missing_cols,
                    "expected_columns": REQUIRED_REQUIREMENT_COLUMNS,
                },
            )

        df["Request-ID"] = df["Request-ID"].astype(str).str.strip()
        df = df[df["Request-ID"].astype(str).str.strip() != ""]
        df = df.dropna(how="all")

        if df.empty:
            raise HTTPException(
                status_code=400,
                detail="Requirement Excel has no valid rows with Request-ID.",
            )

        with open(UPLOADED_REQUIREMENT_PATH, "wb") as f:
            f.write(content)

        return {
            "success": True,
            "message": "Requirement Excel uploaded and validated successfully.",
            "filename": filename,
            "sheet_used": sheet_name,
            "total_jobs": len(df),
        }

    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


@app.post("/start-bulk-analyze")
async def start_bulk_analyze_resumes(
    files: List[UploadFile] = File(...),
    cfg: Settings = Depends(get_settings),
):
    debug("=" * 100)
    debug("Start bulk analyze request received")
    debug(f"Total files received: {len(files)}")

    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    if not os.path.exists(UPLOADED_REQUIREMENT_PATH):
        raise HTTPException(
            status_code=400,
            detail="Please upload valid Requirement Excel before analyzing resumes.",
        )

    if is_active_job_running():
        raise HTTPException(
            status_code=409,
            detail="A resume analysis job is already running. Please wait until it completes.",
        )

    # Delete old status and old uploaded resumes before starting new batch
    reset_bulk_job_storage()

    job_id = f"JOB_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{str(uuid.uuid4())[:8]}"

    saved_files = []

    for index, file in enumerate(files, start=1):
        safe_name = os.path.basename(file.filename or f"resume_{index}.pdf")
        file_path = os.path.join(ACTIVE_UPLOAD_DIR, safe_name)

        content = await file.read()

        with open(file_path, "wb") as f:
            f.write(content)

        saved_files.append(
            {
                "filename": safe_name,
                "content_type": file.content_type or "application/pdf",
                "path": file_path,
            }
        )

    status = get_default_status()
    status.update(
        {
            "job_id": job_id,
            "status": "queued",
            "message": "Resume analysis job queued",
            "total": len(saved_files),
            "processed": 0,
            "successful": 0,
            "failed": 0,
            "skipped": 0,
            "current_file": "",
            "current_batch": f"0/{len(saved_files)}",
            "started_at": datetime.now().isoformat(),
        }
    )

    write_status(status)

    asyncio.create_task(run_bulk_analyze_background_job(job_id, saved_files, cfg))

    return {
        "job_id": job_id,
        "status": "queued",
        "total": len(saved_files),
        "message": "Resume analysis started",
    }


@app.get("/bulk-status")
def get_bulk_status():
    return read_status()


@app.post("/reset-bulk-status")
def reset_bulk_status():
    if is_active_job_running():
        raise HTTPException(
            status_code=409,
            detail="Cannot reset while resume analysis is running.",
        )

    reset_bulk_job_storage()

    status = get_default_status()
    write_status(status)

    return {
        "success": True,
        "message": "Bulk status reset successfully.",
    }


async def run_bulk_analyze_background_job(
    job_id: str,
    saved_files: list,
    cfg: Settings,
):
    start_time = time.time()

    output_rows = []
    tracker_rows = []

    total_input_tokens = 0
    total_output_tokens = 0
    total_tokens = 0

    failed_files = []
    skipped_files = []

    try:
        total_files = len(saved_files)

        # Mark job as processing
        status = read_status()
        status.update(
            {
                "status": "processing",
                "message": "Processing resumes",
                "processed": 0,
                "successful": 0,
                "failed": 0,
                "skipped": 0,
                "current_file": "",
                "current_batch": f"0/{total_files}",
                "resume_logs": [],
                "runtime_info": {
                    "started_at": status.get("started_at", datetime.now().isoformat()),
                    "completed_at": "",
                    "total_seconds": 0,
                    "total_time_text": "",
                    "average_seconds_per_resume": 0,
                },
                "cost_info": calculate_cost_info(
                    input_tokens=0,
                    output_tokens=0,
                    processed_count=0,
                    model_name=cfg.claude_model,
                ),
            }
        )
        write_status(status)

        # Load requirement Excel once
        requirements_df = load_requirement_df(cfg)

        # Process resumes one by one
        for index, item in enumerate(saved_files, start=1):
            filename = item["filename"]

            resume_start_time = time.time()
            resume_started_at = datetime.now().isoformat()

            # Update current processing status before resume starts
            status = read_status()
            status.update(
                {
                    "status": "processing",
                    "message": f"Processing resume {index} of {total_files}",
                    "current_file": filename,
                    "current_batch": f"{index}/{total_files}",
                }
            )
            write_status(status)

            fake_file = SavedUploadFile(
                filename=filename,
                content_type=item.get("content_type") or "application/pdf",
                path=item["path"],
            )

            try:
                result = await process_single_resume_file(
                    file=fake_file,
                    requirements_df=requirements_df,
                    cfg=cfg,
                    index=index,
                    total_files=total_files,
                )

                resume_end_time = time.time()
                resume_seconds = round(resume_end_time - resume_start_time, 2)

                usage = result.get("token_usage", {})

                resume_input_tokens = int(usage.get("input_tokens", 0))
                resume_output_tokens = int(usage.get("output_tokens", 0))
                resume_total_tokens = int(usage.get("total_tokens", 0))

                total_input_tokens += resume_input_tokens
                total_output_tokens += resume_output_tokens
                total_tokens += resume_total_tokens

                resume_log = {
                    "index": index,
                    "filename": result.get("filename", filename),
                    "status": result.get("status", "processed"),
                    "started_at": resume_started_at,
                    "completed_at": datetime.now().isoformat(),
                    "duration_seconds": resume_seconds,
                    "duration_text": format_duration(resume_seconds),
                    "input_tokens": resume_input_tokens,
                    "output_tokens": resume_output_tokens,
                    "total_tokens": resume_total_tokens,
                    "matched_requirements_count": result.get("matched_requirements_count", 0),
                }

                if result["status"] == "skipped":
                    skipped_files.append(result["filename"])

                    status = read_status()

                    resume_logs = status.get("resume_logs", [])
                    resume_logs.append(resume_log)

                    processed_for_cost = max(len(output_rows), 1)

                    status.update(
                        {
                            "status": "processing",
                            "message": f"Skipped resume {index} of {total_files}",
                            "processed": index,
                            "successful": len(output_rows),
                            "failed": len(failed_files),
                            "skipped": len(skipped_files),
                            "current_file": filename,
                            "current_batch": f"{index}/{total_files}",
                            "skipped_files": skipped_files,
                            "failed_files": failed_files,
                            "resume_logs": resume_logs[-200:],
                            "token_usage": {
                                "input_tokens": total_input_tokens,
                                "output_tokens": total_output_tokens,
                                "total_tokens": total_tokens,
                            },
                            "cost_info": calculate_cost_info(
                                input_tokens=total_input_tokens,
                                output_tokens=total_output_tokens,
                                processed_count=processed_for_cost,
                                model_name=cfg.claude_model,
                            ),
                        }
                    )

                    write_status(status)
                    continue

                result_output_rows = result.get("output_rows")
                result_tracker_rows = result.get("tracker_rows")

                if result_output_rows is not None:
                    output_rows.extend(result_output_rows)
                elif result.get("output_row"):
                    output_rows.append(result["output_row"])

                if result_tracker_rows is not None:
                    tracker_rows.extend(result_tracker_rows)
                elif result.get("tracker_row"):
                    tracker_rows.append(result["tracker_row"])

                status = read_status()

                resume_logs = status.get("resume_logs", [])
                resume_logs.append(resume_log)

                processed_for_cost = max(len(output_rows), 1)

                status.update(
                    {
                        "status": "processing",
                        "message": f"Processed resume {index} of {total_files}",
                        "processed": index,
                        "successful": len(output_rows),
                        "failed": len(failed_files),
                        "skipped": len(skipped_files),
                        "current_file": filename,
                        "current_batch": f"{index}/{total_files}",
                        "skipped_files": skipped_files,
                        "failed_files": failed_files,
                        "resume_logs": resume_logs[-200:],
                        "token_usage": {
                            "input_tokens": total_input_tokens,
                            "output_tokens": total_output_tokens,
                            "total_tokens": total_tokens,
                        },
                        "cost_info": calculate_cost_info(
                            input_tokens=total_input_tokens,
                            output_tokens=total_output_tokens,
                            processed_count=processed_for_cost,
                            model_name=cfg.claude_model,
                        ),
                    }
                )

                write_status(status)

            except Exception as e:
                debug(f"Resume failed inside background job: {filename}: {repr(e)}")

                resume_end_time = time.time()
                resume_seconds = round(resume_end_time - resume_start_time, 2)

                failed_item = {
                    "filename": filename,
                    "error": str(e),
                }

                failed_files.append(failed_item)

                resume_log = {
                    "index": index,
                    "filename": filename,
                    "status": "failed",
                    "started_at": resume_started_at,
                    "completed_at": datetime.now().isoformat(),
                    "duration_seconds": resume_seconds,
                    "duration_text": format_duration(resume_seconds),
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "total_tokens": 0,
                    "error": str(e),
                }

                status = read_status()

                resume_logs = status.get("resume_logs", [])
                resume_logs.append(resume_log)

                processed_for_cost = max(len(output_rows), 1)

                status.update(
                    {
                        "status": "processing",
                        "message": f"Resume failed: {filename}. Continuing next resume.",
                        "processed": index,
                        "successful": len(output_rows),
                        "failed": len(failed_files),
                        "skipped": len(skipped_files),
                        "current_file": filename,
                        "current_batch": f"{index}/{total_files}",
                        "failed_files": failed_files,
                        "skipped_files": skipped_files,
                        "resume_logs": resume_logs[-200:],
                        "token_usage": {
                            "input_tokens": total_input_tokens,
                            "output_tokens": total_output_tokens,
                            "total_tokens": total_tokens,
                        },
                        "cost_info": calculate_cost_info(
                            input_tokens=total_input_tokens,
                            output_tokens=total_output_tokens,
                            processed_count=processed_for_cost,
                            model_name=cfg.claude_model,
                        ),
                    }
                )

                write_status(status)

        # If no output rows created, mark full job as failed
        if not output_rows:
            total_time = round(time.time() - start_time, 2)

            status = read_status()
            status.update(
                {
                    "status": "failed",
                    "message": "No valid PDF resumes were processed.",
                    "current_file": "",
                    "current_batch": f"{total_files}/{total_files}",
                    "completed_at": datetime.now().isoformat(),
                    "runtime_info": {
                        "started_at": status.get("started_at", ""),
                        "completed_at": datetime.now().isoformat(),
                        "total_seconds": total_time,
                        "total_time_text": format_duration(total_time),
                        "average_seconds_per_resume": (
                            round(total_time / total_files, 2) if total_files > 0 else 0
                        ),
                    },
                    "cost_info": calculate_cost_info(
                        input_tokens=total_input_tokens,
                        output_tokens=total_output_tokens,
                        processed_count=0,
                        model_name=cfg.claude_model,
                    ),
                }
            )
            write_status(status)
            return

        # Create final Excel
        output_filename = f"{POLLING_OUTPUT_PREFIX}_{job_id}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        create_final_excel(
            output_rows=output_rows,
            tracker_rows=tracker_rows,
            output_path=output_path,
        )

        total_time = round(time.time() - start_time, 2)

        average_seconds_per_resume = 0
        if total_files > 0:
            average_seconds_per_resume = round(total_time / total_files, 2)

        completed_at = datetime.now().isoformat()

        # Mark final completed status
        status = read_status()
        status.update(
            {
                "status": "completed",
                "message": "Bulk resume analysis completed successfully.",
                "processed": total_files,
                "successful": len(output_rows),
                "failed": len(failed_files),
                "skipped": len(skipped_files),
                "current_file": "",
                "current_batch": f"{total_files}/{total_files}",
                "output_filename": output_filename,
                "download_url": f"/download/{output_filename}",
                "failed_files": failed_files,
                "skipped_files": skipped_files,
                "processing_time_seconds": total_time,
                "token_usage": {
                    "input_tokens": total_input_tokens,
                    "output_tokens": total_output_tokens,
                    "total_tokens": total_tokens,
                },
                "cost_info": calculate_cost_info(
                    input_tokens=total_input_tokens,
                    output_tokens=total_output_tokens,
                    processed_count=len(output_rows),
                    model_name=cfg.claude_model,
                ),
                "runtime_info": {
                    "started_at": status.get("started_at", ""),
                    "completed_at": completed_at,
                    "total_seconds": total_time,
                    "total_time_text": format_duration(total_time),
                    "average_seconds_per_resume": average_seconds_per_resume,
                },
                "completed_at": completed_at,
            }
        )

        write_status(status)

    except Exception as e:
        debug(f"Background bulk job failed: {repr(e)}")

        total_time = round(time.time() - start_time, 2)
        completed_at = datetime.now().isoformat()

        status = read_status()
        status.update(
            {
                "status": "failed",
                "message": str(e),
                "current_file": "",
                "completed_at": completed_at,
                "runtime_info": {
                    "started_at": status.get("started_at", ""),
                    "completed_at": completed_at,
                    "total_seconds": total_time,
                    "total_time_text": format_duration(total_time),
                    "average_seconds_per_resume": 0,
                },
                "cost_info": calculate_cost_info(
                    input_tokens=total_input_tokens,
                    output_tokens=total_output_tokens,
                    processed_count=max(len(output_rows), 1),
                    model_name=cfg.claude_model,
                ),
            }
        )

        write_status(status)


@app.get("/jobs")
def list_jobs(cfg: Settings = Depends(get_settings)):
    if not os.path.exists(UPLOADED_REQUIREMENT_PATH):
        return {
            "requirement_uploaded": False,
            "total_jobs": 0,
            "jobs": [],
            "message": "Requirement Excel not uploaded.",
        }

    df = load_requirement_df(cfg)

    jobs = []

    for _, row in df.iterrows():
        row = row.fillna("").to_dict()

        jobs.append(
            {
                "request_id": clean_cell(row.get("Request-ID")),
                "msp_owner": clean_cell(row.get("MSP Owner")),
                "job_title": clean_cell(row.get("Job Title")),
                "skills_name": clean_cell(row.get("Skills - Name")),
                "skills_experience": clean_cell(row.get("Skills - Experience")),
                "status": clean_cell(row.get("Status")),
                "work_location_cdf": clean_cell(row.get("Work Location CDF")),
                "rate_card": clean_cell(row.get("Rate Card")),
                "annually": clean_cell(row.get("Yearly Rate")),
            }
        )

    return {
        "requirement_uploaded": True,
        "total_jobs": len(jobs),
        "jobs": jobs,
    }


@app.get("/jobs/{request_id}")
def get_job(request_id: str, cfg: Settings = Depends(get_settings)):
    requirement = get_requirement_by_id(request_id, cfg)

    return {
        "request_id": request_id,
        "requirement": requirement,
    }

async def process_single_resume_file(
    file: UploadFile,
    requirements_df: pd.DataFrame,
    cfg: Settings,
    index: int,
    total_files: int,
):
    debug("-" * 100)
    debug(f"Processing file {index}/{total_files}: {file.filename}")
    debug(f"Content type: {file.content_type}")

    zero_token_usage = {
        "input_tokens": 0,
        "output_tokens": 0,
        "total_tokens": 0,
    }

    if not is_supported_resume_upload(file):
        return {
            "status": "skipped",
            "filename": file.filename or "unsupported_file",
            "output_rows": [],
            "tracker_rows": [],
            "output_row": None,
            "tracker_row": None,
            "token_usage": zero_token_usage,
        }

    original_file_bytes = await file.read()

    if not original_file_bytes:
        return {
            "status": "skipped",
            "filename": file.filename or "empty_file",
            "output_rows": [],
            "tracker_rows": [],
            "output_row": None,
            "tracker_row": None,
            "token_usage": zero_token_usage,
        }

    if len(original_file_bytes) > cfg.max_pdf_size_mb * 1024 * 1024:
        return {
            "status": "skipped",
            "filename": file.filename or "large_file",
            "output_rows": [],
            "tracker_rows": [],
            "output_row": None,
            "tracker_row": None,
            "token_usage": zero_token_usage,
        }

    try:
        pdf_bytes = get_pdf_bytes_from_resume_upload(
            file_bytes=original_file_bytes,
            original_filename=file.filename or "resume.pdf",
        )
    except Exception as e:
        debug(f"Resume conversion failed for {file.filename}: {repr(e)}")

        error_output_row = {col: "" for col in OUTPUT_COLUMNS}
        error_output_row["CV File Name"] = file.filename or "resume"
        error_output_row["Candidate Name"] = file.filename or "resume"
        error_output_row["ATS"] = "0"
        error_output_row["Experience Mismatch"] = "No"
        error_output_row["Skill Mismatch"] = "No"
        error_output_row["Remark"] = f"Resume conversion failed: {str(e)}"

        tracker_row = {
            "Date": datetime.now().strftime("%d-%m-%Y"),
            "Request ID": "",
            "Status": "",
            "Skills": "",
            "Candidate": file.filename or "resume",
            "Verdict": "Not Suitable",
            "Call Status": "",
            "Remarks": "Not Suitable",
        }

        return {
            "status": "processed",
            "filename": file.filename or "resume",
            "output_rows": [error_output_row],
            "tracker_rows": [tracker_row],
            "output_row": error_output_row,
            "tracker_row": tracker_row,
            "token_usage": zero_token_usage,
        }
    try:
        # ------------------------------------------------------------
        # STEP 1: Extract candidate information from resume using Claude
        # ------------------------------------------------------------
        candidate_info = await asyncio.to_thread(
            call_claude_extract_candidate,
            pdf_bytes,
            file.filename or "resume.pdf",
            cfg,
        )

        # ------------------------------------------------------------
        # STEP 2: Candidate location safety cleanup
        # ------------------------------------------------------------
        if is_likely_company_location(candidate_info):
            candidate_info["candidate_location"] = None
            candidate_info["candidate_city"] = None
            candidate_info["candidate_state"] = None
            candidate_info["candidate_country"] = None

        location_confidence = clean_cell(
            candidate_info.get("candidate_location_confidence")
        ).lower()

        if location_confidence in ["low", "not_found", ""]:
            candidate_info["candidate_location"] = None
            candidate_info["candidate_city"] = None
            candidate_info["candidate_state"] = None
            candidate_info["candidate_country"] = None

        # ------------------------------------------------------------
        # STEP 3: Python shortlist top 25 requirements
        # ------------------------------------------------------------
        shortlisted_requirements = shortlist_requirements(
            requirements_df=requirements_df,
            candidate_info=candidate_info,
            top_n=25,
        )

        # ------------------------------------------------------------
        # STEP 4: Claude returns all relevant requirement matches
        # ------------------------------------------------------------
        match_result = await asyncio.to_thread(
            call_claude_relevant_requirement_matches,
            candidate_info,
            shortlisted_requirements,
            cfg,
        )

        candidate_usage = candidate_info.get("_token_usage", {})
        match_usage = match_result.get("_token_usage", {})

        token_usage = {
            "input_tokens": int(candidate_usage.get("input_tokens", 0))
            + int(match_usage.get("input_tokens", 0)),
            "output_tokens": int(candidate_usage.get("output_tokens", 0))
            + int(match_usage.get("output_tokens", 0)),
            "total_tokens": int(candidate_usage.get("total_tokens", 0))
            + int(match_usage.get("total_tokens", 0)),
        }

        raw_matches = match_result.get("matches", [])

        if not isinstance(raw_matches, list):
            raw_matches = []

        # ------------------------------------------------------------
        # STEP 5: Apply business rule
        # - show all relevant matches
        # - if nothing relevant, show only closest/best one
        # ------------------------------------------------------------
        selected_matches = select_relevant_matches_or_best_one(raw_matches)

        output_rows = []
        tracker_rows = []

        # ------------------------------------------------------------
        # STEP 6: Build rows for selected matches only
        # ------------------------------------------------------------
        for match in selected_matches:
            request_id = (
                clean_cell(match.get("best_request_id"))
                or clean_cell(match.get("request_id"))
                or clean_cell(match.get("Request-ID"))
            )

            req = find_requirement_by_request_id(
                requirements_df=requirements_df,
                request_id=request_id,
            )

            if not req:
                continue

            # Keep compatibility with existing row builders
            match["best_request_id"] = clean_cell(req.get("Request-ID"))
            match["call_status"] = ""

            output_rows.append(
                build_output_row(
                    req=req,
                    candidate_info=candidate_info,
                    match_result=match,
                    filename=file.filename or "resume.pdf",
                )
            )

            tracker_rows.append(
                build_tracker_row(
                    req=req,
                    candidate_info=candidate_info,
                    match_result=match,
                    filename=file.filename or "resume.pdf",
                )
            )

        # ------------------------------------------------------------
        # STEP 7: Fallback if Claude returned nothing usable
        # ------------------------------------------------------------
        if not output_rows and shortlisted_requirements:
            req = shortlisted_requirements[0]

            fallback_match = {
                "best_request_id": clean_cell(req.get("Request-ID")),
                "ats_score": 0,
                "verdict": "Not Suitable",
                "call_status": "",
                "final_remark": "Not Suitable",
                "experience_mismatch": "No",
                "skill_mismatch": "No",
                "location_mismatch": "Not Evaluated",
                "matching_skills": [],
                "missing_skills": [],
                "reason": (
                    "No relevant requirement match returned by AI. "
                    "Closest requirement selected as fallback, but candidate marked Not Suitable."
                ),
            }

            output_rows.append(
                build_output_row(
                    req=req,
                    candidate_info=candidate_info,
                    match_result=fallback_match,
                    filename=file.filename or "resume.pdf",
                )
            )

            tracker_rows.append(
                build_tracker_row(
                    req=req,
                    candidate_info=candidate_info,
                    match_result=fallback_match,
                    filename=file.filename or "resume.pdf",
                )
            )

        return {
            "status": "processed",
            "filename": file.filename or "resume.pdf",
            "output_rows": output_rows,
            "tracker_rows": tracker_rows,
            # compatibility keys
            "output_row": output_rows[0] if output_rows else None,
            "tracker_row": tracker_rows[0] if tracker_rows else None,
            "token_usage": token_usage,
        }

    except Exception as e:
        debug(f"Error processing file {file.filename}: {repr(e)}")

        candidate_name = file.filename or "unknown_resume"

        error_output_row = {col: "" for col in OUTPUT_COLUMNS}
        error_output_row["CV File Name"] = file.filename or "resume.pdf"
        error_output_row["Candidate Name"] = candidate_name
        error_output_row["ATS"] = "0"
        error_output_row["Experience Mismatch"] = "No"
        error_output_row["Skill Mismatch"] = "No"
        error_output_row["Remark"] = f"Error while processing resume: {str(e)}"

        tracker_row = {
            "Date": datetime.now().strftime("%d-%m-%Y"),
            "Request ID": "",
            "Status": "",
            "Skills": "",
            "Candidate": candidate_name,
            "Verdict": "Not Suitable",
            "Call Status": "",
            "Remarks": "Not Suitable",
        }

        return {
            "status": "processed",
            "filename": candidate_name,
            "output_rows": [error_output_row],
            "tracker_rows": [tracker_row],
            "output_row": error_output_row,
            "tracker_row": tracker_row,
            "token_usage": zero_token_usage,
        }



@app.post("/bulk-analyze", response_model=BulkAnalyzeResponse)
async def bulk_analyze_resumes(
    files: List[UploadFile] = File(...),
    cfg: Settings = Depends(get_settings),
):
    start_time = time.time()

    debug("=" * 100)
    debug("Bulk analyze request received")
    debug(f"Total files received: {len(files)}")

    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    if not os.path.exists(UPLOADED_REQUIREMENT_PATH):
        raise HTTPException(
            status_code=400,
            detail="Please upload valid Requirement Excel before analyzing resumes.",
        )

    requirements_df = load_requirement_df(cfg)

    if requirements_df.empty:
        raise HTTPException(status_code=400, detail="Requirement sheet has no rows")

    # Keep one-by-one for Claude rate-limit safety.
    semaphore = asyncio.Semaphore(1)

    async def limited_process(file, index):
        async with semaphore:
            return await process_single_resume_file(
                file=file,
                requirements_df=requirements_df,
                cfg=cfg,
                index=index,
                total_files=len(files),
            )

    tasks = [limited_process(file, index) for index, file in enumerate(files, start=1)]

    results = await asyncio.gather(*tasks)

    output_rows = []
    tracker_rows = []
    skipped_files = []
    processed_count = 0

    total_input_tokens = 0
    total_output_tokens = 0
    total_tokens = 0

    for result in results:
        usage = result.get("token_usage", {})

        total_input_tokens += int(usage.get("input_tokens", 0))
        total_output_tokens += int(usage.get("output_tokens", 0))
        total_tokens += int(usage.get("total_tokens", 0))

        if result["status"] == "skipped":
            skipped_files.append(result["filename"])
            continue

        # IMPORTANT:
        # New max-fit logic returns multiple output rows per resume.
        # So we must use extend(), not append().
        result_output_rows = result.get("output_rows")
        result_tracker_rows = result.get("tracker_rows")

        if result_output_rows is not None:
            output_rows.extend(result_output_rows)
        elif result.get("output_row"):
            output_rows.append(result["output_row"])

        if result_tracker_rows is not None:
            tracker_rows.extend(result_tracker_rows)
        elif result.get("tracker_row"):
            tracker_rows.append(result["tracker_row"])

        processed_count += 1

    if not output_rows:
        raise HTTPException(
            status_code=400,
            detail="No valid PDF resumes were processed",
        )

    file_id = str(uuid.uuid4())[:8]
    output_filename = f"resume_requirement_output_{file_id}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)

    create_final_excel(
        output_rows=output_rows,
        tracker_rows=tracker_rows,
        output_path=output_path,
    )

    total_time = round(time.time() - start_time, 2)

    return BulkAnalyzeResponse(
        message="Bulk resume analysis completed",
        total_files_received=len(files),
        total_pdf_processed=processed_count,
        total_output_rows=len(output_rows),
        processing_time_seconds=total_time,
        output_filename=output_filename,
        download_url=f"/download/{output_filename}",
        skipped_files=skipped_files,
        token_usage={
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "total_tokens": total_tokens,
        },
    )

@app.get("/download/{filename}")
def download_output_file(filename: str):
    # Prevent path traversal
    filename = os.path.basename(filename)
    file_path = os.path.join(OUTPUT_DIR, filename)

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )
    
    
    
@app.post("/reset-requirement")
async def reset_requirement():
    try:
        req_path = os.path.join("input", "uploaded_requirement.xlsx")

        if os.path.exists(req_path):
            os.remove(req_path)

        return {
            "status": "success",
            "message": "Requirement sheet cleared"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


async def process_one_resume_for_db_job(
    job_id: str,
    user_id: str,
    resume_path: str,
    requirements_df,
    cfg: Settings,
    index: int,
    total_files: int,
    semaphore: asyncio.Semaphore,
):
    async with semaphore:
        filename = os.path.basename(resume_path)
        resume_start_time = time.time()

        await mark_resume_processing(job_id, filename)

        await update_job_progress(
            job_id=job_id,
            current_file=filename,
            message=f"Processing resume {index} of {total_files}",
        )

        fake_file = SavedUploadFile(
            filename=filename,
            content_type="application/pdf",
            path=resume_path,
        )

        try:
            result = await process_single_resume_file(
                file=fake_file,
                requirements_df=requirements_df,
                cfg=cfg,
                index=index,
                total_files=total_files,
            )

            resume_seconds = round(time.time() - resume_start_time, 2)

            usage = result.get("token_usage", {})
            resume_input_tokens = int(usage.get("input_tokens", 0))
            resume_output_tokens = int(usage.get("output_tokens", 0))
            resume_total_tokens = int(usage.get("total_tokens", 0))

            result_output_rows = result.get("output_rows") or []
            result_tracker_rows = result.get("tracker_rows") or []

            matched_count = len(result_output_rows)

            if result.get("status") == "skipped":
                await mark_resume_failed(
                    job_id=job_id,
                    filename=filename,
                    duration_seconds=resume_seconds,
                    error_message="Skipped unsupported or empty file",
                )

                await update_job_progress(
                    job_id=job_id,
                    processed_delta=1,
                    skipped_delta=1,
                    input_tokens_delta=resume_input_tokens,
                    output_tokens_delta=resume_output_tokens,
                    total_tokens_delta=resume_total_tokens,
                    current_file=filename,
                    message=f"Skipped resume {index} of {total_files}",
                )

                return {
                    "filename": filename,
                    "status": "skipped",
                    "output_rows": [],
                    "tracker_rows": [],
                    "input_tokens": resume_input_tokens,
                    "output_tokens": resume_output_tokens,
                    "total_tokens": resume_total_tokens,
                    "error": "",
                }

            await save_job_results(
                job_id=job_id,
                user_id=user_id,
                resume_filename=filename,
                output_rows=result_output_rows,
            )

            await mark_resume_completed(
                job_id=job_id,
                filename=filename,
                duration_seconds=resume_seconds,
                input_tokens=resume_input_tokens,
                output_tokens=resume_output_tokens,
                total_tokens=resume_total_tokens,
                matched_requirements_count=matched_count,
            )

            await update_job_progress(
                job_id=job_id,
                processed_delta=1,
                successful_delta=1,
                input_tokens_delta=resume_input_tokens,
                output_tokens_delta=resume_output_tokens,
                total_tokens_delta=resume_total_tokens,
                current_file=filename,
                message=f"Processed resume {index} of {total_files}",
            )

            return {
                "filename": filename,
                "status": "completed",
                "output_rows": result_output_rows,
                "tracker_rows": result_tracker_rows,
                "input_tokens": resume_input_tokens,
                "output_tokens": resume_output_tokens,
                "total_tokens": resume_total_tokens,
                "error": "",
            }

        except Exception as e:
            resume_seconds = round(time.time() - resume_start_time, 2)

            await mark_resume_failed(
                job_id=job_id,
                filename=filename,
                duration_seconds=resume_seconds,
                error_message=str(e),
            )

            await update_job_progress(
                job_id=job_id,
                processed_delta=1,
                failed_delta=1,
                current_file=filename,
                message=f"Failed resume {index} of {total_files}",
            )

            return {
                "filename": filename,
                "status": "failed",
                "output_rows": [],
                "tracker_rows": [],
                "input_tokens": 0,
                "output_tokens": 0,
                "total_tokens": 0,
                "error": str(e),
            }


async def process_db_backed_main_job_background(
    job_id: str,
    user_id: str,
    cfg: Settings,
):
    try:
        batch_start_time = time.time()
        job = await get_user_job_by_id(user_id=user_id, job_id=job_id)
        

        if not job:
            await mark_job_failed(job_id, "Job not found")
            return

        requirement_dir = job.get("requirement_dir")
        resumes_dir = job.get("resumes_dir")
        outputs_dir = job.get("outputs_dir")

        if not requirement_dir or not os.path.exists(requirement_dir):
            await mark_job_failed(job_id, "Requirement folder not found")
            return

        if not resumes_dir or not os.path.exists(resumes_dir):
            await mark_job_failed(job_id, "Resumes folder not found")
            return

        requirement_files = []
        requirement_files.extend(glob.glob(os.path.join(requirement_dir, "*.xlsx")))
        requirement_files.extend(glob.glob(os.path.join(requirement_dir, "*.xls")))

        if not requirement_files:
            await mark_job_failed(job_id, "No requirement Excel found for this job")
            return

        requirement_path = requirement_files[0]

        # Case-insensitive folder scan.
        # This keeps the original processing behavior while supporting .PDF/.DOCX uppercase files.
        resume_files = []
        allowed_resume_extensions_lower = {
            ext.lower() for ext in ALLOWED_RESUME_EXTENSIONS
        }

        for filename in os.listdir(resumes_dir):
            full_path = os.path.join(resumes_dir, filename)

            if not os.path.isfile(full_path):
                continue

            _, ext = os.path.splitext(filename)

            if ext.lower() in allowed_resume_extensions_lower:
                resume_files.append(full_path)

        resume_files = sorted(resume_files)

        if not resume_files:
            await mark_job_failed(job_id, "No resumes found for this job")
            return

        parallel_limit = int(os.getenv("MAIN_AI_PARALLEL_RESUMES", "5"))
        parallel_limit = max(1, min(parallel_limit, 50))

        await mark_job_processing(job_id)

        await update_job_progress(
            job_id=job_id,
            current_file="",
            message=f"Parallel processing started with concurrency={parallel_limit}",
        )

        requirements_df = load_requirement_df_from_path(requirement_path)

        total_files = len(resume_files)
        semaphore = asyncio.Semaphore(parallel_limit)

        tasks = [
            process_one_resume_for_db_job(
                job_id=job_id,
                user_id=user_id,
                resume_path=resume_path,
                requirements_df=requirements_df,
                cfg=cfg,
                index=index,
                total_files=total_files,
                semaphore=semaphore,
            )
            for index, resume_path in enumerate(resume_files, start=1)
        ]

        task_results = await asyncio.gather(*tasks)

        output_rows = []
        tracker_rows = []

        total_input_tokens = 0
        total_output_tokens = 0
        total_tokens = 0

        failed_files = []
        skipped_files = []

        for item in task_results:
            output_rows.extend(item.get("output_rows") or [])
            tracker_rows.extend(item.get("tracker_rows") or [])

            total_input_tokens += int(item.get("input_tokens") or 0)
            total_output_tokens += int(item.get("output_tokens") or 0)
            total_tokens += int(item.get("total_tokens") or 0)

            if item.get("status") == "failed":
                failed_files.append(item.get("filename"))
            elif item.get("status") == "skipped":
                skipped_files.append(item.get("filename"))

        if not output_rows:
            await mark_job_failed(job_id, "No valid output rows generated")
            return

        output_filename = f"resume_requirement_output_{job_id}.xlsx"
        output_path = os.path.join(outputs_dir, output_filename)

        create_final_excel(
            output_rows=output_rows,
            tracker_rows=tracker_rows,
            output_path=output_path,
        )

        # await mark_job_completed(job_id, output_file_path=output_path)

        batch_seconds = round(time.time() - batch_start_time, 2)
        average_seconds_per_resume = 0

        if total_files > 0:
            average_seconds_per_resume = round(batch_seconds / total_files, 2)

        await mark_job_completed(
            job_id,
            output_file_path=output_path,
            processing_time_seconds=batch_seconds,
            average_seconds_per_resume=average_seconds_per_resume,
        )

        await update_job_progress(
            job_id=job_id,
            current_file="",
            message=f"Job completed successfully. Failed={len(failed_files)}, Skipped={len(skipped_files)}",
        )

    except Exception as e:
        await mark_job_failed(job_id, str(e))
        
@app.post("/jobs/{job_id}/start")
async def start_db_backed_main_job(
    job_id: str,
    background_tasks: BackgroundTasks,
    cfg: Settings = Depends(get_settings),
    current_user: dict = Depends(get_current_user),
):
    user_id = str(current_user["_id"])

    job = await get_user_job_by_id(user_id=user_id, job_id=job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    if job.get("mode") != "main_ai":
        raise HTTPException(
            status_code=400,
            detail="This endpoint currently supports only main_ai jobs",
        )

    if job.get("status") == "processing":
        raise HTTPException(status_code=409, detail="Job is already processing")

    if job.get("status") == "completed":
        raise HTTPException(status_code=409, detail="Job is already completed")

    await update_job_progress(
        job_id=job_id,
        message="Job accepted for background processing",
    )

    background_tasks.add_task(
        process_db_backed_main_job_background,
        job_id,
        user_id,
        cfg,
    )

    return {
        "success": True,
        "message": "Job processing started in background",
        "job_id": job_id,
        "status": "processing_started",
    }
